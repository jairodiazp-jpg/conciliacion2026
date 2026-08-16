from __future__ import annotations

import base64
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================================================
# COLOR ADQUIRENCIAS
# ============================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="FF00FF00",
)


class AdquirenciasConciliador:

    # ========================================================
    # ENCABEZADOS ESPERADOS
    # ========================================================

    ADQ_AUTH_HEADER = "CODIGO AUTORIZACION"
    ADQ_VALUE_HEADER = "VALOR TOTAL"
    ADQ_OBS_HEADER = "Observacion"

    CCS_AUTH_HEADER = "N° De Aprobación"
    CCS_VALUE_HEADER = "Valor"
    CCS_OBS_HEADER = "Observaciones"

    def __init__(
        self,
        adquirencias_bytes: bytes,
        ccs_bytes: bytes,
    ) -> None:

        self.adquirencias_wb = load_workbook(
            filename=BytesIO(adquirencias_bytes)
        )

        self.ccs_wb = load_workbook(
            filename=BytesIO(ccs_bytes)
        )

    # ========================================================
    # DETECTAR COLUMNAS Y FILAS DINÁMICAMENTE
    # ========================================================

    def _detectar_estructura(self, ws, headers: list[str]) -> dict[str, int]:
        """Busca encabezados en las primeras 30 filas."""
        found = {}
        for fila in range(1, min(ws.max_row, 30) + 1):
            for cell in ws[fila]:
                if cell.value is None: continue
                val = str(cell.value).strip()
                for h in headers:
                    if val.lower() == h.lower():
                        found[h] = cell.column
            if len(found) == len(headers):
                break
        return found

    def _detectar_inicio_datos(self, ws, auth_col: int, *, fallback: int = 19) -> int:
        """Busca la primera fila después del encabezado que tenga datos."""
        encabezados_auth = {
            "n° de aprobación",
            "n de aprobacion",
            "codigo autorizacion",
            "código autorización",
        }
        for fila in range(1, min(ws.max_row, 30) + 5):
            val = ws.cell(row=fila, column=auth_col).value
            if val is not None and str(val).strip().lower() in encabezados_auth:
                return fila + 1
        return fallback

    # ========================================================
    # NORMALIZAR AUTORIZACIÓN
    # ========================================================

    def _normalizar_auth(self, value: Any) -> str:

        if value is None:
            return ""

        text = str(value)

        text = (
            text
            .replace("\u00A0", " ")
            .replace("\u200B", "")
            .replace("\u200C", "")
            .replace("\u200D", "")
            .replace("\ufeff", "")
        )

        return text.strip()

    # ========================================================
    # NORMALIZAR VALOR
    # ========================================================

    def _normalizar_valor(
        self,
        value: Any,
    ) -> Decimal:

        if value is None:
            return Decimal("0.00")

        if isinstance(value, Decimal):

            return value.quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        if isinstance(value, int):

            return Decimal(value).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        if isinstance(value, float):

            return Decimal(str(value)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        text = str(value).strip()

        if not text:
            return Decimal("0.00")

        text = (
            text
            .replace("$", "")
            .replace("COP", "")
            .replace("cop", "")
            .replace("\u00A0", "")
            .replace(" ", "")
        )

        negativo = False

        if text.startswith("(") and text.endswith(")"):

            negativo = True
            text = text[1:-1]

        if text.startswith("-"):

            negativo = True
            text = text[1:]

        # ---------------------------------------------
        # FORMATO 1.250.000,50
        # ---------------------------------------------

        if "," in text and "." in text:

            if text.rfind(",") > text.rfind("."):

                text = text.replace(".", "")
                text = text.replace(",", ".")

            else:

                text = text.replace(",", "")

        # ---------------------------------------------
        # FORMATO 1250000,50
        # ---------------------------------------------

        elif "," in text:

            partes = text.split(",")

            if len(partes[-1]) <= 2:

                text = (
                    "".join(partes[:-1])
                    + "."
                    + partes[-1]
                )

            else:

                text = "".join(partes)

        # ---------------------------------------------
        # FORMATO 1.250.000
        # ---------------------------------------------

        elif "." in text:

            partes = text.split(".")

            if len(partes) > 2:

                text = "".join(partes)

        try:

            resultado = Decimal(text)

            if negativo:
                resultado = -resultado

            return resultado.quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        except (InvalidOperation, ValueError):

            return Decimal("0.00")

    # ========================================================
    # VALOR EN CENTAVOS
    # ========================================================

    def _valor_key(
        self,
        value: Any,
    ) -> int:

        valor = self._normalizar_valor(value)

        return int(
            (
                valor * Decimal("100")
            ).quantize(
                Decimal("1"),
                rounding=ROUND_HALF_UP,
            )
        )

    # ========================================================
    # BUSCAR HOJA CCS
    # ========================================================

    def _buscar_hoja_ccs(self):

        objetivo = "bancolombia cta 690"

        for ws in self.ccs_wb.worksheets:

            nombre = (
                str(ws.title)
                .strip()
                .lower()
            )

            if nombre == objetivo:

                return ws

        # Fallback seguro
        for ws in self.ccs_wb.worksheets:

            nombre = (
                str(ws.title)
                .strip()
                .lower()
            )

            if (
                "bancolombia" in nombre
                and "690" in nombre
            ):

                return ws

        raise ValueError(
            "No se encontró la hoja "
            "'Bancolombia Cta 690'. "
            f"Hojas disponibles: {self.ccs_wb.sheetnames}"
        )

    # ========================================================
    # DETECTAR COLUMNAS CCS
    # ========================================================

    def _detectar_columnas_ccs(self, ws) -> tuple[int, int]:
        return self.CCS_AUTH_COL, self.CCS_VALUE_COL

    # ========================================================
    # BUSCAR COLUMNA POR ENCABEZADO
    # ========================================================

    def _buscar_columna(
        self,
        ws,
        nombres: set[str],
        filas: tuple[int, ...],
    ) -> int | None:

        for fila in filas:

            if fila > ws.max_row:
                continue

            for cell in ws[fila]:

                if cell.value is None:
                    continue

                nombre = (
                    str(cell.value)
                    .strip()
                    .lower()
                )

                if nombre in nombres:

                    return cell.column

        return None

    # ========================================================
    # ENCONTRAR COLUMNA OBSERVACIONES ADQUIRENCIAS
    # ========================================================

    def _obtener_columna_observaciones_adq(
        self,
        ws,
    ) -> int:

        nombres = {
            "observacion",
            "observaciones",
        }

        # Primero buscar por encabezado
        for fila in range(1, min(ws.max_row, 10) + 1):

            for cell in ws[fila]:

                if cell.value is None:
                    continue

                texto = (
                    str(cell.value)
                    .strip()
                    .lower()
                )

                if texto in nombres:

                    return cell.column

        # Si no existe, crearla al final
        nueva_columna = ws.max_column + 1

        ws.cell(
            row=1,
            column=nueva_columna,
        ).value = "OBSERVACIONES"

        return nueva_columna

    # ========================================================
    # ENCONTRAR COLUMNA OBSERVACIONES ADQUIRENCIAS (AMPLIADO)
    # ========================================================

    def _obtener_columna_observaciones_adq_ampliado(
        self,
        ws,
    ) -> int:

        nombres = {
            "observacion",
            "observaciones",
        }

        # Buscar en las primeras 30 filas
        for fila in range(1, min(ws.max_row, 30) + 1):

            for cell in ws[fila]:

                if cell.value is None:
                    continue

                texto = (
                    str(cell.value)
                    .strip()
                    .lower()
                )

                if texto in nombres:

                    return cell.column

        # Si no existe, crearla al final
        nueva_columna = ws.max_column + 1

        ws.cell(
            row=1,
            column=nueva_columna,
        ).value = "OBSERVACIONES"

        return nueva_columna

    # ========================================================
    # ENCONTRAR COLUMNA OBSERVACIONES CCS
    # ========================================================

    def _obtener_columna_observaciones_ccs(
        self,
        ws,
        ccs_auth_col: int,
        ccs_val_col: int,
    ) -> tuple[int, int]:

        nombres = {
            "observacion",
            "observaciones",
        }

        # Buscar en las filas donde normalmente
        # se encuentra el encabezado.
        for fila in range(
            1,
            min(ws.max_row, 30) + 1,
        ):

            for cell in ws[fila]:

                if cell.value is None:
                    continue

                texto = (
                    str(cell.value)
                    .strip()
                    .lower()
                )

                if texto in nombres:

                    return cell.column, fila

        # Si no existe, crearla al final
        columna = ws.max_column + 1

        # Buscar la fila donde están Auth y Valor
        header_row = 1

        for fila in range(
            1,
            min(ws.max_row, 30) + 1,
        ):

            auth_header = str(
                ws.cell(
                    row=fila,
                    column=ccs_auth_col,
                ).value
                or ""
            ).strip().lower()

            value_header = str(
                ws.cell(
                    row=fila,
                    column=ccs_val_col,
                ).value
                or ""
            ).strip().lower()

            if (
                "aprob" in auth_header
                or "autoriz" in auth_header
            ) and "valor" in value_header:

                header_row = fila
                break

        ws.cell(
            row=header_row,
            column=columna,
        ).value = "OBSERVACIONES"

        return columna, header_row

    # ========================================================
    # ENCONTRAR PRIMERA FILA DE DATOS CCS
    # ========================================================

    def _obtener_inicio_ccs(
        self,
        ws,
        ccs_auth_col: int,
        ccs_val_col: int,
    ) -> int:
        return self.CCS_START_ROW

    # ========================================================
    # AGREGAR OBSERVACIÓN
    # ========================================================

    def _agregar_observacion(
        self,
        ws,
        row: int,
        col: int,
        texto: str,
    ) -> None:

        actual = str(
            ws.cell(
                row=row,
                column=col,
            ).value
            or ""
        ).strip()

        if not actual:

            ws.cell(
                row=row,
                column=col,
            ).value = texto

            return

        if texto not in actual:

            ws.cell(
                row=row,
                column=col,
            ).value = (
                actual
                + " | "
                + texto
            )

    # ========================================================
    # PROCESAR
    # ========================================================

    def procesar(self) -> dict[str, Any]:

        # ====================================================
        # HOJAS
        # ====================================================

        adq_sheet = self.adquirencias_wb.active

        ccs_sheet = self._buscar_hoja_ccs()

        # ====================================================
        # COLUMNAS
        # ====================================================

        adq_struct = self._detectar_estructura(adq_sheet, [self.ADQ_AUTH_HEADER, self.ADQ_VALUE_HEADER, self.ADQ_OBS_HEADER])
        adq_auth_col = adq_struct.get(self.ADQ_AUTH_HEADER, 23)
        adq_val_col = adq_struct.get(self.ADQ_VALUE_HEADER, 16)
        
        # Usar el método ampliado para buscar la columna de observaciones
        adq_obs_col = self._obtener_columna_observaciones_adq_ampliado(adq_sheet)

        ccs_struct = self._detectar_estructura(ccs_sheet, [self.CCS_AUTH_HEADER, self.CCS_VALUE_HEADER, self.CCS_OBS_HEADER])
        ccs_auth_col = ccs_struct.get(self.CCS_AUTH_HEADER, 6)
        ccs_val_col = ccs_struct.get(self.CCS_VALUE_HEADER, 8)
        ccs_obs_col, ccs_header_row = self._obtener_columna_observaciones_ccs(ccs_sheet, ccs_auth_col, ccs_val_col)

        # ====================================================
        # INICIO DATOS CCS
        # ====================================================

        ccs_start_row = self._detectar_inicio_datos(ccs_sheet, ccs_auth_col)
        adq_start_row = self._detectar_inicio_datos(adq_sheet, adq_auth_col, fallback=2)

        # ====================================================
        # ÍNDICE CCS
        # ====================================================

        ccs_index: dict[
            tuple[str, int],
            list[int],
        ] = {}

        cantidad_ccs = 0

        for row in range(
            ccs_start_row,
            ccs_sheet.max_row + 1,
        ):

            auth = self._normalizar_auth(
                ccs_sheet.cell(
                    row=row,
                    column=ccs_auth_col,
                ).value
            )

            if not auth:
                continue

            raw_value = ccs_sheet.cell(
                row=row,
                column=ccs_val_col,
            ).value

            value_key = self._valor_key(
                raw_value
            )

            key = (
                auth,
                value_key,
            )

            ccs_index.setdefault(
                key,
                [],
            ).append(row)

            cantidad_ccs += 1

        # ====================================================
        # PROCESAR ADQUIRENCIAS
        # ====================================================

        cruce_count = 0
        cantidad_adq = 0
        cantidad_auth = 0

        for row in range(
            adq_start_row,
            adq_sheet.max_row + 1,
        ):

            cantidad_adq += 1

            auth = self._normalizar_auth(
                adq_sheet.cell(
                    row=row,
                    column=adq_auth_col,
                ).value
            )

            if not auth:
                continue

            cantidad_auth += 1

            raw_value = adq_sheet.cell(
                row=row,
                column=adq_val_col,
            ).value

            value = self._normalizar_valor(
                raw_value
            )

            value_key = self._valor_key(
                raw_value
            )

            # =================================================
            # CRUCE REAL
            #
            # AUTORIZACIÓN + VALOR
            # =================================================

            key = (
                auth,
                value_key,
            )

            matches = ccs_index.get(
                key,
                [],
            )

            if not matches:
                continue

            # =================================================
            # NUEVA ADQUIRENCIA
            # =================================================

            cruce_count += 1

            adquirencia_nombre = (
                f"ADQUIRENCIA {cruce_count}"
            )

            # =================================================
            # PINTAR ADQUIRENCIA
            # =================================================

            for cell in adq_sheet[row]:

                cell.fill = GREEN_FILL

            # =================================================
            # OBSERVACIÓN ADQUIRENCIA
            # =================================================

            filas_ccs = ", ".join(
                str(x)
                for x in matches
            )

            observacion_adq = (
                f"{adquirencia_nombre} cruza con CCS Memorando (hoja {ccs_sheet.title.strip()}) | "
                f"Fila(s) CCS: {filas_ccs} | "
                f"Autorización: {auth} | "
                f"Valor: {value:.2f}"
            )

            self._agregar_observacion(
                adq_sheet,
                row,
                adq_obs_col,
                observacion_adq,
            )

            # =================================================
            # MARCAR CCS
            # =================================================

            for ccs_row in matches:

                # ---------------------------------------------
                # Pintar la fila correspondiente en CCS
                # ---------------------------------------------

                for cell in ccs_sheet[ccs_row]:

                    cell.fill = GREEN_FILL

                # ---------------------------------------------
                # Observación CCS
                # ---------------------------------------------

                observacion_ccs = (
                    f"{adquirencia_nombre} cruza con Adquirencias fila {row} | "
                    f"Autorización: {auth} | "
                    f"Valor: {value:.2f}"
                )

                self._agregar_observacion(
                    ccs_sheet,
                    ccs_row,
                    ccs_obs_col,
                    observacion_ccs,
                )

        # ====================================================
        # GUARDAR ADQUIRENCIAS
        # ====================================================

        out_adq = BytesIO()

        self.adquirencias_wb.save(
            out_adq
        )

        adquirencias_result = (
            out_adq.getvalue()
        )

        # ====================================================
        # GUARDAR CCS
        # ====================================================

        out_ccs = BytesIO()

        self.ccs_wb.save(
            out_ccs
        )

        ccs_result = (
            out_ccs.getvalue()
        )

        # ====================================================
        # VERIFICACIÓN DEL ARCHIVO FINAL
        # ====================================================

        # Volver a abrir los archivos guardados para
        # garantizar que las modificaciones realmente
        # quedaron persistidas.

        verificacion_adq = load_workbook(
            filename=BytesIO(
                adquirencias_result
            )
        )

        verificacion_ccs = load_workbook(
            filename=BytesIO(
                ccs_result
            )
        )

        ver_adq_sheet = (
            verificacion_adq.active
        )

        ver_ccs_sheet = (
            verificacion_ccs[
                ccs_sheet.title
            ]
        )

        observaciones_adq_final = 0
        filas_verdes_adq = 0

        for row in range(
            adq_start_row,
            ver_adq_sheet.max_row + 1,
        ):

            obs = (
                ver_adq_sheet.cell(
                    row=row,
                    column=adq_obs_col,
                ).value
            )

            if obs and "ADQUIRENCIA" in str(obs):

                observaciones_adq_final += 1

                # Revisar si alguna celda de la fila
                # tiene el verde configurado.
                for cell in ver_adq_sheet[row]:

                    if (
                        cell.fill
                        and cell.fill.fill_type == "solid"
                        and cell.fill.fgColor.rgb
                        and cell.fill.fgColor.rgb.upper()
                        in {
                            "FF00FF00",
                            "0000FF00",
                        }
                    ):

                        filas_verdes_adq += 1
                        break

        observaciones_ccs_final = 0
        filas_verdes_ccs = 0

        for row in range(
            ccs_start_row,
            ver_ccs_sheet.max_row + 1,
        ):

            obs = (
                ver_ccs_sheet.cell(
                    row=row,
                    column=ccs_obs_col,
                ).value
            )

            if obs and "ADQUIRENCIA" in str(obs):

                observaciones_ccs_final += 1

                for cell in ver_ccs_sheet[row]:

                    if (
                        cell.fill
                        and cell.fill.fill_type == "solid"
                        and cell.fill.fgColor.rgb
                        and cell.fill.fgColor.rgb.upper()
                        in {
                            "FF00FF00",
                            "0000FF00",
                        }
                    ):

                        filas_verdes_ccs += 1
                        break

        # ====================================================
        # RESULTADO
        # ====================================================

        return {

            "adquirencias_file": (
                base64.b64encode(
                    adquirencias_result
                ).decode("utf-8")
            ),

            "ccs_file": (
                base64.b64encode(
                    ccs_result
                ).decode("utf-8")
            ),

            "resumen": {

                "registros_adquirencias":
                    cantidad_adq,

                "autorizaciones_adquirencias":
                    cantidad_auth,

                "registros_ccs":
                    cantidad_ccs,

                "cruzados":
                    cruce_count,

                "observaciones_adquirencias":
                    observaciones_adq_final,

                "filas_verdes_adquirencias":
                    filas_verdes_adq,

                "observaciones_ccs":
                    observaciones_ccs_final,

                "filas_verdes_ccs":
                    filas_verdes_ccs,

                "hoja_ccs":
                    ccs_sheet.title,

                "criterio_cruce":
                    "AUTORIZACION + VALOR EXACTO",

                "adquirencias_autorizacion":
                    "W",

                "adquirencias_valor":
                    "P",

                "ccs_autorizacion":
                    "F",

                "ccs_valor":
                    "H",
            },
        }