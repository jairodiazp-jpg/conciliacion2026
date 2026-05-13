"""Verificar qué filas se anotaron."""

import base64
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from procesador_adquirencias import ProcesadorAdquirencias

base = Path('..') / 'tmp_e2e'
adq_bytes = (base / 'ADQUIRENCIAS MAYO (1).xlsx').read_bytes()
memo_bytes = (base / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx').read_bytes()
memo_b64 = base64.b64encode(memo_bytes).decode('utf-8')

processor = ProcesadorAdquirencias(adq_bytes, memo_b64)
adq_data = processor._extraer_adquirencias_con_fila()
processor._cruzar_ambos_archivos(adq_data)

print(f"Filas anotadas (primeras 10): {sorted(processor._annotated_rows)[:10]}")
print(f"Total filas anotadas: {len(processor._annotated_rows)}")

# Verificar esas filas específicas
sheet_690 = processor.contable_workbook['Bancolombia Cta 690 ']
if processor._annotated_rows:
    first_annotated = sorted(processor._annotated_rows)[0]
    print(f"\nPrimera fila anotada: {first_annotated}")
    print(f"  col 17 value: '{sheet_690.cell(first_annotated, 17).value}'")
    print(f"  col 18 value: '{sheet_690.cell(first_annotated, 18).value}'")
    print(f"  col 1 value: '{sheet_690.cell(first_annotated, 1).value}'")
