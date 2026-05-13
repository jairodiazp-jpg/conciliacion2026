"""Verificar si las columnas 17 y 18 están fusionadas en la hoja 690."""

from pathlib import Path
from openpyxl import load_workbook

# Cargar original
memo_path = Path('..') / 'tmp_e2e' / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx'
wb = load_workbook(memo_path)

for sheet in wb.worksheets:
    if '690' in sheet.title:
        print(f"Hoja: {sheet.title}\n")
        print(f"Celdas fusionadas (merged cells): {sheet.merged_cells}")
        
        # Verificar si hay MergedCells en columnas 17 y 18
        merged_ranges = list(sheet.merged_cells.ranges)
        print(f"Total merged ranges: {len(merged_ranges)}\n")
        
        for merged_range in merged_ranges:
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            
            # Verificar si afecta columnas 17 o 18
            if (min_col <= 17 <= max_col) or (min_col <= 18 <= max_col):
                print(f"Merged: {merged_range} (cols {min_col}-{max_col}, rows {min_row}-{max_row})")
        
        # Intentar encontrar una fila donde las columnas 17 y 18 no estén fusionadas
        print("\nVerificando filas 2-50:")
        for row in range(2, 51):
            cell_17 = sheet.cell(row, 17)
            cell_18 = sheet.cell(row, 18)
            is_merged_17 = isinstance(cell_17.value, type(None)) and str(type(cell_17).__name__) == 'MergedCell'
            is_merged_18 = isinstance(cell_18.value, type(None)) and str(type(cell_18).__name__) == 'MergedCell'
            if row <= 10:
                print(f"  Row {row}: col17={type(cell_17).__name__}, col18={type(cell_18).__name__}")
        break
