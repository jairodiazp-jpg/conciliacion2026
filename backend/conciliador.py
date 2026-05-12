from __future__ import annotations

import base64
import itertools
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from validacion_temporal import (
    RangoFechas,
    ValidacionTemporalConfig,
    calcular_rango_fechas,
    evaluar_temporal,
    inferir_periodo_principal,
)


UPPER_MARKER = "sin registrar en libros"
LOWER_MARKER = "sin registrar en el extracto"

ORANGE_FILL = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


@dataclass
class LedgerEntry:
    sheet_name: str
    row: int
    value_col: int
    value: float
    section: str
    raw_date: date | None
    approval_number: str | None
    number: str | None
    matched: bool = False


class ConciliadorContable:
    def __init__(self, file_bytes: bytes, temporal_config: ValidacionTemporalConfig | None = None) -> None:
        self.workbook = load_workbook(filename=BytesIO(file_bytes))
        self.logs: list[dict[str, Any]] = []
        self.cross_id = 1
        self.possible_id = 1
        self.temporal_config = temporal_config or ValidacionTemporalConfig()
        self.rango_operativo: RangoFechas | None = None
        self.annotation_columns = self._index_annotation_columns()

    def procesar(self) -> dict[str, Any]:
        entries = self._extraer_movimientos()
        self.total_entries = len(entries)
        self._configurar_rango_temporal(entries)

        upper_entries = [entry for entry in entries if entry.section == "upper"]
        lower_entries = [entry for entry in entries if entry.section == "lower"]

        self._fase_aprobacion(upper_entries, lower_entries)
        self._fase_valor_fecha(upper_entries, lower_entries)
        self._fase_uno_a_muchos(upper_entries, lower_entries)
        self._fase_posibles(upper_entries, lower_entries)

        resumen = self._construir_resumen()
        alertas = self._construir_alertas(resumen)

        output_stream = BytesIO()
        self.workbook.save(output_stream)
        excel_b64 = base64.b64encode(output_stream.getvalue()).decode("utf-8")

        return {
            "file": excel_b64,
            "logs": self.logs,
            "resumen": resumen,
            "alertas": alertas,
        }

    def _normalizar_texto(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value)
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _configurar_rango_temporal(self, entries: list[LedgerEntry]) -> None:
        fechas = [entry.raw_date for entry in entries if entry.raw_date is not None]
        rango_general = calcular_rango_fechas(fechas)
        periodo_principal = inferir_periodo_principal(fechas)

        if periodo_principal is None:
            self.rango_operativo = rango_general
            return

        fechas_principales = [
            value
            for value in fechas
            if value is not None and (value.year, value.month) == periodo_principal
        ]
        if fechas_principales and len(fechas_principales) / max(1, len(fechas)) >= 0.6:
            self.rango_operativo = calcular_rango_fechas(fechas_principales)
            self._append_log(
                tipo="info_temporal",
                valor=0.0,
                fecha=self.rango_operativo.fecha_minima,
                confianza=1.0,
                detalle=(
                    f"Periodo principal detectado: {periodo_principal[0]}-{periodo_principal[1]:02d}. "
                    f"Rango operativo {self.rango_operativo.fecha_minima} a {self.rango_operativo.fecha_maxima}"
                ),
            )
            return

        self.rango_operativo = rango_general

    def _index_annotation_columns(self) -> dict[str, tuple[int | None, int | None]]:
        indexed: dict[str, tuple[int | None, int | None]] = {}
        for sheet in self.workbook.worksheets:
            comment_col: int | None = None
            observation_col: int | None = None

            max_scan_row = min(sheet.max_row, 25)
            for row in sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)

                    # Detectar columnas de comentario/observacion de forma tolerante
                    if comment_col is None and "coment" in header:
                        comment_col = cell.column
                    if observation_col is None and "observ" in header:
                        observation_col = cell.column

                if comment_col is not None and observation_col is not None:
                    break

            indexed[sheet.title] = (comment_col, observation_col)

        return indexed

    def _append_text_once(self, current_value: Any, text: str) -> str:
        current = str(current_value).strip() if current_value is not None else ""
        if not current:
            return text
        if text.lower() in current.lower():
            return current
        return f"{current} | {text}"

    def _account_label(self, sheet_name: str) -> str:
        match = re.search(r"\d{4,}", sheet_name)
        if match:
            return match.group(0)
        return sheet_name

    def _comment_account_text(self, source_sheet_name: str, target_sheet_name: str) -> str:
        source_account = self._account_label(source_sheet_name)
        targets = [item.strip() for item in target_sheet_name.split(",") if item.strip()]
        if not targets:
            targets = [target_sheet_name]

        target_accounts = sorted({self._account_label(item) for item in targets})
        if len(target_accounts) == 1 and target_accounts[0] == source_account:
            return f"Cruce en la misma cuenta {source_account}"
        if len(target_accounts) == 1:
            return f"Cruce de cuenta {source_account} a cuenta {target_accounts[0]}"
        return f"Cruce de cuenta {source_account} a cuentas {', '.join(target_accounts)}"

    def _annotate_entry(self, entry: LedgerEntry, other: LedgerEntry | str, tag: str | None = None) -> None:
        """Anota una fila con comentario y observacion.

        - `other` puede ser otro LedgerEntry (mejor) o el nombre de la hoja destino.
        - `tag` si se proporciona (ej. 'Posible Cruce 1' o 'Cruzado 2') se incluye en el comentario
          y se asegura que el homolog reciba la misma enumeración y una traza "Encontrado en ...".
        """
        sheet = self.workbook[entry.sheet_name]
        comment_col, observation_col = self.annotation_columns.get(entry.sheet_name, (None, None))

        # Resolver información del otro extremo para traza
        if isinstance(other, LedgerEntry):
            other_sheet_name = other.sheet_name
            other_location = f"{other.sheet_name}:fila {other.row}"
        else:
            other_sheet_name = str(other)
            other_location = other_sheet_name

        # Texto base para cuentas / reclasificacion
        base_text = self._comment_account_text(entry.sheet_name, other_sheet_name)

        if comment_col is not None:
            comment_cell = sheet.cell(row=entry.row, column=comment_col)
            if tag:
                comment_text = f"{tag} - {base_text}"
            else:
                comment_text = f"Posible Cruce - {base_text}"
            comment_cell.value = self._append_text_once(comment_cell.value, comment_text)

        if observation_col is not None:
            observation_cell = sheet.cell(row=entry.row, column=observation_col)
            # Incluir trazabilidad explícita indicando dónde se encontró el homolog
            if tag:
                observation_text = f"{tag} encontrado en {other_location}. {base_text}"
            else:
                observation_text = f"Encontrado en {other_location}. Reclasificación: {base_text}"
            observation_cell.value = self._append_text_once(observation_cell.value, observation_text)

    def _annotate_pair(self, left: LedgerEntry, right: LedgerEntry, tag: str | None = None) -> None:
        # Anotación recíproca consistente; pasa el objeto para traza completa
        self._annotate_entry(left, right, tag)
        self._annotate_entry(right, left, tag)

    def _annotate_group(self, primary: LedgerEntry, related_entries: list[LedgerEntry], tag: str | None = None) -> None:
        related_sheets = sorted({entry.sheet_name for entry in related_entries})
        relation_target = ", ".join(related_sheets) if related_sheets else primary.sheet_name
        # Anotar el primario indicando los relacionados
        if tag:
            self._annotate_entry(primary, relation_target, tag)
        else:
            self._annotate_entry(primary, relation_target)
        for entry in related_entries:
            self._annotate_entry(entry, primary, tag)

    def _parse_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(cleaned, fmt).date()
                    except Exception:
                        continue
        return None

    def _parse_number(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, str):
            cleaned = value.strip()
            if len(cleaned) >= 3:
                return cleaned
        return None

    def _parse_value_candidates(self, row_cells: list[Any], sheet: Worksheet) -> list[tuple[int, float]]:
        candidates: list[tuple[int, float]] = []
        for idx, cell in enumerate(row_cells, start=1):
            value = cell.value
            if value is None or isinstance(value, bool):
                continue
            if getattr(cell, "is_date", False):
                continue
            if isinstance(value, (int, float)):
                candidates.append((idx, float(value)))
        return candidates

    def _extraer_movimientos(self) -> list[LedgerEntry]:
        entries: list[LedgerEntry] = []

        for sheet in self.workbook.worksheets:
            current_section: str | None = None

            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                texts = [
                    self._normalizar_texto(str(cell.value))
                    for cell in row
                    if isinstance(cell.value, str) and str(cell.value).strip()
                ]

                if any(UPPER_MARKER in text for text in texts):
                    current_section = "upper"
                    continue
                if any(LOWER_MARKER in text for text in texts):
                    current_section = "lower"
                    continue

                if current_section is None:
                    continue

                value_candidates = self._parse_value_candidates(list(row), sheet)
                if not value_candidates:
                    continue

                value_col, value = max(value_candidates, key=lambda item: abs(item[1]))
                if abs(value) < 0.0001:
                    continue

                parsed_date: date | None = None
                parsed_approval_number: str | None = None
                parsed_number: str | None = None

                if len(row) >= 6:
                    parsed_approval_number = self._parse_number(row[5].value)

                for cell in row:
                    if parsed_date is None:
                        parsed_date = self._parse_date(cell.value)
                    if parsed_number is None:
                        parsed_number = self._parse_number(cell.value)

                entries.append(
                    LedgerEntry(
                        sheet_name=sheet.title,
                        row=row[0].row,
                        value_col=value_col,
                        value=value,
                        section=current_section,
                        raw_date=parsed_date,
                        approval_number=parsed_approval_number,
                        number=parsed_number,
                    )
                )

        return entries

    def _tag_and_color(self, entry: LedgerEntry, color: PatternFill, tag: str) -> None:
        sheet = self.workbook[entry.sheet_name]
        value_cell = sheet.cell(row=entry.row, column=entry.value_col)
        value_cell.fill = color

        tag_cell = sheet.cell(row=entry.row, column=entry.value_col + 1)
        current = str(tag_cell.value).strip() if tag_cell.value is not None else ""
        if current:
            if tag not in current:
                tag_cell.value = f"{current} | {tag}"
        else:
            tag_cell.value = tag

    def _append_log(self, tipo: str, valor: float, fecha: date | None, confianza: float, detalle: str) -> None:
        self.logs.append(
            {
                "tipo": tipo,
                "valor": round(valor, 2),
                "fecha": fecha.isoformat() if fecha else None,
                "confianza": round(confianza, 2),
                "detalle": detalle,
            }
        )

    def _approval_key(self, entry: LedgerEntry) -> str | None:
        # Prioriza el numero de aprobacion de columna F y conserva fallback al numero detectado.
        return entry.approval_number or entry.number

    def _same_abs_value(self, left: float, right: float) -> bool:
        return abs(abs(left) - abs(right)) <= 0.0001

    def _date_gap_days(self, left: date | None, right: date | None) -> int | None:
        if left is None or right is None:
            return None
        return abs((left - right).days)

    def _temporal_evaluation(self, left: date | None, right: date | None) -> tuple[bool, int | None, str, int | None]:
        decision = evaluar_temporal(left, right, self.temporal_config, self.rango_operativo)
        return decision.permitida, decision.prioridad, decision.motivo, decision.diferencia_dias

    def _log_temporal_rejection(self, left: LedgerEntry, right: LedgerEntry, motivo: str) -> None:
        self._append_log(
            tipo="rechazado_temporal",
            valor=left.value,
            fecha=left.raw_date or right.raw_date,
            confianza=0.0,
            detalle=(
                f"Movimiento rechazado: Fecha PSE: {left.raw_date.isoformat() if left.raw_date else None}. "
                f"Fecha contable: {right.raw_date.isoformat() if right.raw_date else None}. Motivo: {motivo}"
            ),
        )

    def _has_same_approval(self, upper: LedgerEntry, lower: LedgerEntry) -> bool:
        upper_approval = self._approval_key(upper)
        lower_approval = self._approval_key(lower)
        return bool(upper_approval and lower_approval and upper_approval == lower_approval)

    def _is_confirmed_pair(self, upper: LedgerEntry, lower: LedgerEntry) -> bool:
        permitida, _, _, _ = self._temporal_evaluation(upper.raw_date, lower.raw_date)
        if not permitida:
            return False
        if not self._has_same_approval(upper, lower):
            return False
        if not self._same_abs_value(upper.value, lower.value):
            return False
        date_gap = self._date_gap_days(upper.raw_date, lower.raw_date)
        return date_gap is not None and date_gap <= 3

    def _fase_aprobacion(self, uppers: list[LedgerEntry], lowers: list[LedgerEntry]) -> None:
        lower_by_number: dict[str, list[LedgerEntry]] = {}
        for lower in lowers:
            approval_key = self._approval_key(lower)
            if lower.matched or not approval_key:
                continue
            lower_by_number.setdefault(approval_key, []).append(lower)

        for upper in uppers:
            approval_key = self._approval_key(upper)
            if upper.matched or not approval_key:
                continue
            candidates = lower_by_number.get(approval_key, [])
            ranked_candidates: list[tuple[int, int, LedgerEntry]] = []
            for lower in candidates:
                if lower.matched:
                    continue
                permitida, prioridad, motivo, gap = self._temporal_evaluation(upper.raw_date, lower.raw_date)
                if not permitida:
                    if self._has_same_approval(upper, lower) and self._same_abs_value(upper.value, lower.value):
                        self._log_temporal_rejection(upper, lower, motivo)
                    continue
                if not self._is_confirmed_pair(upper, lower):
                    continue

                ranked_candidates.append((prioridad or 99, gap or 999, lower))

            candidate = min(ranked_candidates, default=None, key=lambda item: (item[0], item[1], item[2].row))
            if candidate is None:
                continue
            best_lower = candidate[2]

            tag = f"Cruzado Aprob {self.cross_id}"
            upper.matched = True
            best_lower.matched = True
            self._tag_and_color(upper, ORANGE_FILL, tag)
            self._tag_and_color(best_lower, ORANGE_FILL, tag)
            self._annotate_pair(upper, best_lower, tag)

            gap_text = self._date_gap_days(upper.raw_date, best_lower.raw_date)
            gap_detail = f" y fecha gap {gap_text} dias" if gap_text is not None else ""
            self._append_log(
                tipo="aprobacion",
                valor=upper.value,
                fecha=upper.raw_date or best_lower.raw_date,
                confianza=0.99,
                detalle=(
                    f"Coincidencia de numero de aprobacion {approval_key}, valor y fecha{gap_detail} entre "
                    f"{upper.sheet_name}:fila {upper.row} y {best_lower.sheet_name}:fila {best_lower.row}"
                ),
            )
            self.cross_id += 1

    def _fase_valor_fecha(self, uppers: list[LedgerEntry], lowers: list[LedgerEntry]) -> None:
        for upper in uppers:
            if upper.matched or upper.raw_date is None:
                continue

            best_candidate: LedgerEntry | None = None
            best_rank: tuple[int, int, int] | None = None

            for lower in lowers:
                if lower.matched or lower.raw_date is None:
                    continue
                permitida, prioridad, motivo, gap = self._temporal_evaluation(upper.raw_date, lower.raw_date)
                if not permitida:
                    if self._same_abs_value(upper.value, lower.value):
                        self._log_temporal_rejection(upper, lower, motivo)
                    continue
                if not self._is_confirmed_pair(upper, lower):
                    continue

                candidate_rank = (prioridad or 99, gap or 999, lower.row)
                if best_rank is None or candidate_rank < best_rank:
                    best_candidate = lower
                    best_rank = candidate_rank

            if best_candidate is None:
                continue

            tag = f"Cruzado {self.cross_id}"
            upper.matched = True
            best_candidate.matched = True
            self._tag_and_color(upper, GREEN_FILL, tag)
            self._tag_and_color(best_candidate, GREEN_FILL, tag)
            self._annotate_pair(upper, best_candidate, tag)
            self._append_log(
                tipo="valor_fecha",
                valor=upper.value,
                fecha=upper.raw_date,
                confianza=0.9,
                detalle=(
                    f"Coincidencia de aprobacion, valor y fecha (±3 dias) entre "
                    f"{upper.sheet_name}:fila {upper.row} y {best_candidate.sheet_name}:fila {best_candidate.row}"
                ),
            )
            self.cross_id += 1

    def _find_combination(self, target_entry: LedgerEntry, pool: list[LedgerEntry]) -> list[LedgerEntry] | None:
        target = target_entry.value
        target_cents = int(round(abs(target) * 100))
        if target_cents == 0:
            return None

        candidates: list[LedgerEntry] = []
        for entry in pool:
            if entry.matched:
                continue
            permitida, _, motivo, _ = self._temporal_evaluation(target_entry.raw_date, entry.raw_date)
            if not permitida:
                if self._has_same_approval(target_entry, entry):
                    self._log_temporal_rejection(target_entry, entry, motivo)
                continue
            if not self._has_same_approval(target_entry, entry):
                continue
            date_gap = self._date_gap_days(target_entry.raw_date, entry.raw_date)
            if date_gap is None or date_gap > 3:
                continue
            if int(round(abs(entry.value) * 100)) > target_cents:
                continue
            candidates.append(entry)
        candidates = sorted(
            candidates,
            key=lambda item: (
                self._temporal_evaluation(target_entry.raw_date, item.raw_date)[1] or 99,
                self._date_gap_days(target_entry.raw_date, item.raw_date) or 999,
                -abs(item.value),
            ),
        )[:18]

        for size in range(2, 5):
            for combo in itertools.combinations(candidates, size):
                combo_sum = sum(int(round(abs(item.value) * 100)) for item in combo)
                if combo_sum == target_cents:
                    return list(combo)
        return None

    def _fase_uno_a_muchos(self, uppers: list[LedgerEntry], lowers: list[LedgerEntry]) -> None:
        for upper in uppers:
            if upper.matched:
                continue
            combo = self._find_combination(upper, lowers)
            if not combo:
                continue

            tag = f"Cruzado {self.cross_id}"
            upper.matched = True
            self._tag_and_color(upper, GREEN_FILL, tag)
            for item in combo:
                item.matched = True
                self._tag_and_color(item, GREEN_FILL, tag)
            self._annotate_group(upper, combo, tag)

            self._append_log(
                tipo="valor_fecha",
                valor=upper.value,
                fecha=upper.raw_date,
                confianza=0.86,
                detalle=(
                    f"Cruce uno-a-muchos con aprobacion y fecha validada ({len(combo)} movimientos) para "
                    f"{upper.sheet_name}:fila {upper.row}"
                ),
            )
            self.cross_id += 1

        for lower in lowers:
            if lower.matched:
                continue
            combo = self._find_combination(lower, uppers)
            if not combo:
                continue

            tag = f"Cruzado {self.cross_id}"
            lower.matched = True
            self._tag_and_color(lower, GREEN_FILL, tag)
            for item in combo:
                item.matched = True
                self._tag_and_color(item, GREEN_FILL, tag)
            self._annotate_group(lower, combo, tag)

            self._append_log(
                tipo="valor_fecha",
                valor=lower.value,
                fecha=lower.raw_date,
                confianza=0.86,
                detalle=(
                    f"Cruce uno-a-muchos con aprobacion y fecha validada ({len(combo)} movimientos) para "
                    f"{lower.sheet_name}:fila {lower.row}"
                ),
            )
            self.cross_id += 1

    def _fase_posibles(self, uppers: list[LedgerEntry], lowers: list[LedgerEntry]) -> None:
        lower_map: dict[int, list[LedgerEntry]] = {}
        for lower in lowers:
            if lower.matched:
                continue
            key = int(round(abs(lower.value) * 100))
            lower_map.setdefault(key, []).append(lower)

        for upper in uppers:
            if upper.matched:
                continue

            key = int(round(abs(upper.value) * 100))
            candidates = lower_map.get(key, [])
            ranked_candidates: list[tuple[int, int, LedgerEntry]] = []
            for item in candidates:
                if item.matched:
                    continue
                permitida, prioridad, motivo, gap = self._temporal_evaluation(upper.raw_date, item.raw_date)
                if not permitida:
                    self._log_temporal_rejection(upper, item, motivo)
                    continue
                ranked_candidates.append((prioridad or 99, gap or 999, item))
            candidate = min(ranked_candidates, default=None, key=lambda item: (item[0], item[1], item[2].row))
            if candidate is None:
                continue
            candidate = candidate[2]

            tag = f"Posible Cruce {self.possible_id}"
            upper.matched = True
            candidate.matched = True
            self._tag_and_color(upper, YELLOW_FILL, tag)
            self._tag_and_color(candidate, YELLOW_FILL, tag)
            self._annotate_pair(upper, candidate, tag)

            self._append_log(
                tipo="posible",
                valor=upper.value,
                fecha=upper.raw_date or candidate.raw_date,
                confianza=0.55,
                detalle=(
                    f"Mismo valor detectado sin validacion completa entre "
                    f"{upper.sheet_name}:fila {upper.row} y {candidate.sheet_name}:fila {candidate.row}"
                ),
            )
            self.possible_id += 1

    def _construir_resumen(self) -> dict[str, Any]:
        cruzados = len([log for log in self.logs if log["tipo"] in {"aprobacion", "valor_fecha"}])
        posibles = len([log for log in self.logs if log["tipo"] == "posible"])
        total = cruzados + posibles
        precision = (cruzados / total) if total else 0.0

        total_entries = getattr(self, "total_entries", 0)
        porcentaje_cruzados = (cruzados / total_entries * 100) if total_entries else 0.0
        porcentaje_posibles = (posibles / total_entries * 100) if total_entries else 0.0

        return {
            "cruzados": cruzados,
            "posibles": posibles,
            "precision_estimada": round(precision, 2),
            "total_movements": total_entries,
            "porcentaje_cruzados": round(porcentaje_cruzados, 2),
            "porcentaje_posibles": round(porcentaje_posibles, 2),
        }

    def _construir_alertas(self, resumen: dict[str, Any]) -> list[str]:
        alertas: list[str] = []

        cruzados = int(resumen["cruzados"])
        posibles = int(resumen["posibles"])

        if cruzados == 0:
            alertas.append("No se detectaron cruces concluyentes.")
        if posibles >= 3 and posibles > max(cruzados, 1):
            alertas.append("Se detecto una cantidad elevada de posibles cruces. Revisar manualmente.")

        return alertas
