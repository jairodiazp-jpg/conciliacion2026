"""Debug detallado: Inspeccionar TODAS las columnas del memorando 690."""

from pathlib import Path
from openpyxl import load_workbook

def main():
    print("=== ANÁLISIS COMPLETO MEMORANDO 690 ===\n")
    
    memorando_path = Path(__file__).parent.parent / "tmp_e2e" / "CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx"
    
    memo_wb = load_workbook(memorando_path)
    
    # Encontrar hoja con 690
    ws_690 = None
    for sheet_name in memo_wb.sheetnames:
        if '690' in sheet_name:
            ws_690 = memo_wb[sheet_name]
            break
    
    if not ws_690:
        print("✗ NO ENCONTRADA HOJA CON 690")
        return
    
    print(f"Hoja: {ws_690.title}")
    print(f"Máximo fila: {ws_690.max_row}, Máximo columna: {ws_690.max_column}\n")
    
    # Mostrar las primeras 30 filas completamente
    print("PRIMERAS 30 FILAS (todas las columnas):\n")
    for row_idx in range(1, min(31, ws_690.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(ws_690.max_column + 1, 10)):  # Primeras 9 columnas
            cell = ws_690.cell(row=row_idx, column=col_idx)
            value = str(cell.value)[:20] if cell.value else ""
            row_data.append(f"C{col_idx}:{value}")
        print(f"F{row_idx:3d}: {' | '.join(row_data)}")
    
    # Buscar el marcador y analizar estructura
    print(f"\n\nBUSCANDO MARCADOR...")
    for row_idx in range(1, min(30, ws_690.max_row + 1)):
        row_text = " ".join([str(cell.value) for cell in ws_690[row_idx] if cell.value])
        if 'consignaciones' in row_text.lower():
            print(f"Encontrado en fila {row_idx}: {row_text[:80]}")
            marker_row = row_idx
            break
    
    # Después del marcador
    print(f"\nFILAS DESPUÉS DEL MARCADOR (filas {marker_row + 1} a {marker_row + 15}):\n")
    for row_idx in range(marker_row + 1, min(marker_row + 15, ws_690.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(ws_690.max_column + 1, 16)):
            cell = ws_690.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, str):
                row_data.append(f"C{col_idx}:{value[:15]}")
            elif isinstance(value, (int, float)):
                row_data.append(f"C{col_idx}:{value}")
            else:
                row_data.append(f"C{col_idx}:{value}")
        print(f"F{row_idx}: {' | '.join(row_data)}")
    
    # Análisis de tipos
    print(f"\n\nTIPOS DE DATOS EN LAS FILAS CON DATOS:")
    for row_idx in range(marker_row + 5, min(marker_row + 10, ws_690.max_row + 1)):
        print(f"\nFila {row_idx}:")
        for col_idx in range(1, ws_690.max_column + 1):
            cell = ws_690.cell(row=row_idx, column=col_idx)
            print(f"  Col {col_idx}: tipo={type(cell.value).__name__}, valor={cell.value}")

if __name__ == '__main__':
    main()
