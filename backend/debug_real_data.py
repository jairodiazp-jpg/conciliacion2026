"""Debug: Inspeccionar datos reales en ambos archivos."""

from pathlib import Path
from openpyxl import load_workbook

def main():
    print("=== INSPECCIÓN DE DATOS REALES ===\n")
    
    # Rutas
    adquirencias_path = Path(__file__).parent.parent / "tmp_e2e" / "ADQUIRENCIAS MAYO (1).xlsx"
    memorando_path = Path(__file__).parent.parent / "tmp_e2e" / "CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx"
    
    # Adquirencias
    print("1. ARCHIVO ADQUIRENCIAS")
    adq_wb = load_workbook(adquirencias_path)
    for sheet_name in adq_wb.sheetnames:
        ws = adq_wb[sheet_name]
        print(f"\n   Hoja: '{sheet_name}'")
        print(f"   Max fila: {ws.max_row}, Max columna: {ws.max_column}")
        
        # Header
        headers = []
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=1, column=col)
            headers.append(f"Col{col}: {cell.value}")
        print(f"   Headers: {', '.join(headers)}")
        
        # Primeras 5 filas de datos
        print(f"   Primeros datos:")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col)
                row_data.append(str(cell.value)[:15])
            print(f"      Fila {row_idx}: {' | '.join(row_data)}")
    
    # Memorando 690
    print("\n2. ARCHIVO MEMORANDO - Hoja 690")
    memo_wb = load_workbook(memorando_path)
    
    ws_690 = None
    for sheet_name in memo_wb.sheetnames:
        if '690' in sheet_name:
            ws_690 = memo_wb[sheet_name]
            print(f"\n   Hoja encontrada: '{sheet_name}'")
            print(f"   Max fila: {ws_690.max_row}, Max columna: {ws_690.max_column}")
            
            # Headers
            headers = []
            for col in range(1, min(ws_690.max_column + 1, 10)):
                cell = ws_690.cell(row=1, column=col)
                headers.append(f"Col{col}: {cell.value}")
            print(f"   Headers: {', '.join(headers)}")
            
            # Buscar marcador "consignaciones sin registrar"
            print(f"\n   Buscando marcador 'consignaciones sin registrar'...")
            for row_idx in range(1, min(50, ws_690.max_row + 1)):
                row_text = " ".join([str(cell.value) for cell in ws_690[row_idx] if cell.value])
                if 'consignaciones' in row_text.lower():
                    print(f"      ✓ Encontrado en fila {row_idx}")
                    print(f"      Contenido: {row_text[:80]}")
                    start_data_row = row_idx + 1
                    break
            else:
                print(f"      ✗ NO ENCONTRADO")
                start_data_row = 2
            
            # Primeros datos después del marcador
            print(f"\n   Primeros datos (desde fila {start_data_row}):")
            for row_idx in range(start_data_row, min(start_data_row + 5, ws_690.max_row + 1)):
                row_data = []
                for col in range(1, min(7, ws_690.max_column + 1)):
                    cell = ws_690.cell(row=row_idx, column=col)
                    row_data.append(str(cell.value)[:15])
                print(f"      Fila {row_idx}: {' | '.join(row_data)}")
            break
    
    if ws_690 is None:
        print(f"   ✗ NO ENCONTRADA HOJA CON 690")
    
    # Comparación
    print("\n3. ANÁLISIS DE COINCIDENCIAS POTENCIALES")
    print("\n   Comparando primeros valores de Adquirencias con 690...")
    
    # Extraer datos de Adquirencias (columnas conocidas)
    adq_ws = adq_wb["Hoja2"]
    adq_datos = []
    for row_idx in range(2, min(10, adq_ws.max_row + 1)):
        fecha = adq_ws.cell(row=row_idx, column=2).value  # B
        valor = adq_ws.cell(row=row_idx, column=16).value  # P
        auth = adq_ws.cell(row=row_idx, column=23).value  # W
        if valor is not None:
            adq_datos.append((row_idx, fecha, valor, auth))
    
    # Extraer datos de 690
    if ws_690:
        memo_datos = []
        for row_idx in range(2, min(50, ws_690.max_row + 1)):
            # Buscar donde empiezan los datos (después del marcador)
            if ws_690.cell(row=row_idx, column=1).value and 'consignaciones' in str(ws_690.cell(row=row_idx, column=1).value).lower():
                continue
            
            fecha = ws_690.cell(row=row_idx, column=2).value  # B
            valor = ws_690.cell(row=row_idx, column=4).value  # D
            auth = ws_690.cell(row=row_idx, column=6).value  # F
            
            # Solo si hay valor
            if valor and isinstance(valor, (int, float)) and valor > 0:
                memo_datos.append((row_idx, fecha, valor, auth))
                if len(memo_datos) >= 5:
                    break
        
        print(f"\n   ADQUIRENCIAS ({len(adq_datos)} filas):")
        for row, fecha, valor, auth in adq_datos:
            print(f"      F{row}: Fecha={fecha} | Valor={valor} | Auth={auth}")
        
        print(f"\n   MEMORANDO 690 ({len(memo_datos)} filas):")
        for row, fecha, valor, auth in memo_datos:
            print(f"      F{row}: Fecha={fecha} | Valor={valor} | Auth={auth}")
        
        print(f"\n   BÚSQUEDA DE COINCIDENCIAS:")
        for adq_row, adq_fecha, adq_valor, adq_auth in adq_datos:
            for memo_row, memo_fecha, memo_valor, memo_auth in memo_datos:
                if adq_valor == memo_valor and adq_auth == memo_auth:
                    print(f"      ✓ COINCIDENCIA: Valor={adq_valor}, Auth={adq_auth}")
                elif adq_valor == memo_valor:
                    print(f"      ? Valor coincide ({adq_valor}) pero Auth NO: Adq={adq_auth} vs Memo={memo_auth}")
                elif adq_auth == memo_auth:
                    print(f"      ? Auth coincide ({adq_auth}) pero Valor NO: Adq={adq_valor} vs Memo={memo_valor}")

if __name__ == '__main__':
    main()
