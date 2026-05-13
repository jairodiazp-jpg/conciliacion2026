"""Diagnóstico detallado de dónde se crean las columnas de comentarios."""

import base64
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from procesador_adquirencias import ProcesadorAdquirencias

base = Path('..') / 'tmp_e2e'
adq = (base / 'ADQUIRENCIAS MAYO (1).xlsx').read_bytes()
memo = (base / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx').read_bytes()

memo_b64 = base64.b64encode(memo).decode('utf-8')

print("=== DIAGNÓSTICO DE COLUMNAS DE COMENTARIOS ===\n")

# Inspeccionar memorando ANTES de procesarlo
print("1. MEMORANDO ORIGINAL:")
memo_wb_original = load_workbook(Path('..') / 'tmp_e2e' / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx')
for sheet in memo_wb_original.worksheets:
    if '690' in sheet.title.lower():
        print(f"   Hoja: {sheet.title}")
        print(f"   Max column: {sheet.max_column}")
        row1_headers = [f"{col}: {sheet.cell(1, col).value}" for col in range(1, sheet.max_column + 1)]
        for h in row1_headers:
            if 'coment' in str(h).lower() or 'observ' in str(h).lower():
                print(f"      ✓ {h}")
        break

# Procesar
processor = ProcesadorAdquirencias(adq, memo_b64)

print("\n2. DESPUÉS DE _ensure_annotation_columns:")
print(f"   Contable sheets: {[s.title for s in processor.contable_workbook.worksheets]}")
for sheet in processor.contable_workbook.worksheets:
    if '690' in sheet.title.lower():
        print(f"   Hoja: {sheet.title}")
        print(f"   Max column ahora: {sheet.max_column}")
        row1_headers = [f"{col}: {sheet.cell(1, col).value}" for col in range(1, sheet.max_column + 1)]
        for h in row1_headers:
            if 'coment' in str(h).lower() or 'observ' in str(h).lower():
                print(f"      ✓ {h}")

print(f"\n3. annotation_columns_cont (indexadas):")
for sheet_title, (comment_col, obs_col) in processor.annotation_columns_cont.items():
    if '690' in sheet_title.lower():
        print(f"   {sheet_title}: comment={comment_col}, observation={obs_col}")

print("\n✓ DIAGNÓSTICO COMPLETADO")
