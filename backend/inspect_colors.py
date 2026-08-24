import sys
import os
from openpyxl import load_workbook

def inspect_colors():
    files = ["../test_file.xlsx", "../test.xlsx"]
    for file_path in files:
        print(f"\n--- Inspecting colors in {file_path} ---")
        try:
            wb = load_workbook(file_path, data_only=False) # data_only=False to get styles
            for sheet in wb.worksheets:
                print(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(min_row=1, max_row=5):
                    row_colors = []
                    for cell in row:
                        fill = cell.fill
                        color = fill.start_color.index if fill.start_color else None
                        row_colors.append(f"{cell.value}: {color}")
                    print(row_colors)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

if __name__ == "__main__":
    inspect_colors()