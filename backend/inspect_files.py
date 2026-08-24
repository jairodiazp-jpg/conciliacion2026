import sys
import os
from openpyxl import load_workbook

def inspect():
    files = ["../test_file.xlsx", "../test.xlsx"]
    for file_path in files:
        print(f"\n--- Inspecting {file_path} ---")
        try:
            wb = load_workbook(file_path, data_only=True)
            for sheet in wb.worksheets:
                print(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                    print(row)
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

if __name__ == "__main__":
    inspect()