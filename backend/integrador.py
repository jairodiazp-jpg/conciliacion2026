from __future__ import annotations

import base64
import unicodedata
from datetime import date, datetime, timedelta
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from conciliador import ConciliadorContable
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from pse_conciliador import PseConciliador
from agrupacion_pse import conciliacion_por_agrupacion
from procesador_adquirencias import ProcesadorAdquirencias
from validacion_temporal import ValidacionTemporalConfig, evaluar_temporal


@dataclass
class ArchivoResultado:
    name: str
    content: str


class ProcesadorIntegrado:
    def __init__(
        self,
        *,
        contable_bytes: bytes | None = None,
        pse_bytes: bytes | None = None,
        cruces_bytes: bytes | None = None,
        adquirencias_bytes: bytes | None = None,
        date_tolerance_days: int = 1,
        value_tolerance: float = 0.01,
        temporal_config: ValidacionTemporalConfig | None = None,
    ) -> None:
        self.contable_bytes = contable_bytes
        self.pse_bytes = pse_bytes
        self.cruces_bytes = cruces_bytes
        self.adquirencias_bytes = adquirencias_bytes
        self.date_tolerance_days = date_tolerance_days
        self.value_tolerance = value_tolerance
        self.temporal_config = temporal_config or ValidacionTemporalConfig()

    def _normalizar_texto(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value)
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _index_annotation_columns(self, workbook) -> dict[str, tuple[int | None, int | None]]:
        indexed: dict[str, tuple[int | None, int | None]] = {}
        for sheet in workbook.worksheets:
            comment_col: int | None = None
            observation_col: int | None = None

            max_scan_row = min(sheet.max_row, 25)
            for row in sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)

                    if comment_col is None and header in {"comentario", "comentarios"}:
                        comment_col = cell.column
                    if observation_col is None and header in {"observacion", "observaciones"}:
                        observation_col = cell.column

                if comment_col is not None and observation_col is not None:
                    break

            indexed[sheet.title] = (comment_col, observation_col)

        return indexed

    def _ensure_annotation_columns(self, workbook) -> dict[str, tuple[int | None, int | None]]:
        annotation_columns: dict[str, tuple[int | None, int | None]] = {}
        for sheet in workbook.worksheets:
            comment_col: int | None = None
            observation_col: int | None = None
            header_row: int = 1
            max_scan_row = min(sheet.max_row, 25)
            for row in sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)
                    if comment_col is None and header in {"comentario", "comentarios"}:
                        comment_col = cell.column
                    if observation_col is None and header in {"observacion", "observaciones"}:
                        observation_col = cell.column
                if comment_col is not None and observation_col is not None:
                    break
            if comment_col is None:
                comment_col = sheet.max_column + 1
                sheet.cell(row=header_row, column=comment_col).value = "Comentario"
            if observation_col is None:
                observation_col = sheet.max_column + 1
                sheet.cell(row=header_row, column=observation_col).value = "Observacion"
            annotation_columns[sheet.title] = (comment_col, observation_col)
        return annotation_columns

    def _append_text_once(self, current_value: Any, text: str) -> str:
        current = str(current_value).strip() if current_value is not None else ""
        if not current:
            return text
        if text.lower() in current.lower():
            return current
        return f"{current} | {text}"

    def _parse_row_date(self, row) -> datetime | None:
        for cell in row:
            value = cell.value
            if isinstance(value, datetime):
                return value
            if isinstance(value, date):
                return datetime.combine(value, datetime.min.time())
            if isinstance(value, str):
                cleaned = value.strip()
                if cleaned:
                    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S"):
                        try:
                            return datetime.strptime(cleaned, fmt)
                        except Exception:
                            continue
            if getattr(cell, "is_date", False) and isinstance(value, (int, float)):
                try:
                    parsed = from_excel(value)
                    if isinstance(parsed, datetime):
                        return parsed
                    if isinstance(parsed, date):
                        return datetime.combine(parsed, datetime.min.time())
                except Exception:
                    continue
        return None

    def _parse_row_value(self, row) -> float | None:
        best: float | None = None
        for cell in row:
            value = cell.value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                v = float(value)
                if abs(v) < 0.0001:
                    continue
                if best is None or abs(v) > abs(best):
                    best = v
        return best

    def _parse_date_text(self, text: str) -> date | None:
        cleaned = (text or "").strip()
        if not cleaned:
            return None
        # normaliza separadores comunes
        cleaned = cleaned.replace(".", "/")
        # ISO primero
        try:
            return datetime.fromisoformat(cleaned).date()
        except Exception:
            pass
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except Exception:
                continue
        return None

    def _is_bancolombia_sheet(self, sheet_name: str) -> bool:
        name = (sheet_name or "").lower()
        if "bancolombia" in name:
            return True
        if "1331" in name:
            return True
        return False

    def _row_contains_pse_marker(self, row) -> bool:
        texts = [self._normalizar_texto(str(cell.value)) for cell in row if isinstance(cell.value, str)]
        return any("pago virtual pse" in text for text in texts)

    def _annotate_pse_rows_in_contable(self, contable_b64: str, pse_dataset: list[dict[str, Any]]) -> str:
        workbook = load_workbook(filename=BytesIO(base64.b64decode(contable_b64)))
        annotation_columns = self._ensure_annotation_columns(workbook)
        for pse_row in pse_dataset:
            group_id = str(pse_row.get("id_grupo_conciliacion") or "").strip()
            if not group_id:
                continue
            try:
                pse_value = abs(float(pse_row.get("valor")))
            except (TypeError, ValueError):
                continue
            pse_date = str(pse_row.get("fecha") or "").strip()
            if not pse_date:
                continue
            pseudo_comment = str(pse_row.get("comentario_conciliacion") or "").strip()
            pseudo_values = str(pse_row.get("valores_asociados") or "").strip()
            group_text = f"[{group_id}]"
            for sheet in workbook.worksheets:
                comment_col, observation_col = annotation_columns.get(sheet.title, (None, None))
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    row_date = self._parse_row_date(row)
                    if row_date is None:
                        continue
                    row_value = self._parse_row_value(row)
                    if row_value is None:
                        continue
                    row_date_text = row_date.date().isoformat() if isinstance(row_date, datetime) else row_date.isoformat()
                    if row_date_text != pse_date:
                        continue
                    if abs(abs(row_value) - pse_value) > self.value_tolerance:
                        continue
                    if comment_col is not None:
                        comment_cell = sheet.cell(row=row[0].row, column=comment_col)
                        comment_text = f"Cruce PSE {group_text} - {pseudo_comment if pseudo_comment else pseudo_values}"
                        comment_cell.value = self._append_text_once(comment_cell.value, comment_text)
                    if observation_col is not None:
                        observation_cell = sheet.cell(row=row[0].row, column=observation_col)
                        base_observation = pseudo_values if pseudo_values else "Reclasificacion PSE detectada"
                        observation_text = f"{group_text} {base_observation}".strip()
                        observation_cell.value = self._append_text_once(observation_cell.value, observation_text)
        output_stream = BytesIO()
        workbook.save(output_stream)
        return base64.b64encode(output_stream.getvalue()).decode("utf-8")

    def _merge_pse_comments_into_contable(self, contable_b64: str, dataset_cruces: list[dict[str, Any]]) -> str:
        workbook = load_workbook(filename=BytesIO(base64.b64decode(contable_b64)))
        annotation_columns = self._index_annotation_columns(workbook)
        for row in dataset_cruces:
            sheet_name = row.get("sheet")
            excel_row = row.get("row")
            estado = str(row.get("estado_conciliacion") or "").strip()
            comentario = str(row.get("comentario_conciliacion") or "").strip()
            valores = str(row.get("valores_asociados") or "").strip()
            if not sheet_name or sheet_name not in workbook.sheetnames:
                continue
            if estado == "Sin coincidencia":
                continue
            if not comentario and not valores:
                continue
            try:
                excel_row_int = int(excel_row)
            except (TypeError, ValueError):
                continue
            if excel_row_int < 1:
                continue
            sheet = workbook[sheet_name]
            comment_col, observation_col = annotation_columns.get(sheet_name, (None, None))
            group_id = str(row.get("id_grupo_conciliacion") or "").strip()
            group_text = f"[{group_id}]" if group_id else ""
            if comment_col is not None:
                comment_cell = sheet.cell(row=excel_row_int, column=comment_col)
                base_comment = comentario if comentario else valores
                prefix = f"Cruce PSE {group_text}".strip()
                comment_text = f"{prefix} - {base_comment}" if base_comment else prefix
                comment_cell.value = self._append_text_once(comment_cell.value, comment_text)
            if observation_col is not None:
                observation_cell = sheet.cell(row=excel_row_int, column=observation_col)
                base_observation = valores if valores else "Reclasificacion PSE detectada"
                observation_text = f"{group_text} {base_observation}".strip() if group_text else base_observation
                observation_cell.value = self._append_text_once(observation_cell.value, observation_text)
        output_stream = BytesIO()
        workbook.save(output_stream)
        return base64.b64encode(output_stream.getvalue()).decode("utf-8")

    def procesar(self) -> dict[str, Any]:
        contable_result: dict[str, Any] | None = None
        pse_result: dict[str, Any] | None = None
        logs: list[dict[str, Any]] = []
        alertas: list[str] = []
        if self.contable_bytes is not None:
            contable_result = ConciliadorContable(
                self.contable_bytes,
                temporal_config=self.temporal_config,
            ).procesar()
            logs.extend(contable_result.get("logs", []))
            alertas.extend(contable_result.get("alertas", []))
        if self.pse_bytes is not None and self.cruces_bytes is not None:
            pse_result = PseConciliador(
                self.pse_bytes,
                self.cruces_bytes,
                date_tolerance_days=self.date_tolerance_days,
                value_tolerance=self.value_tolerance,
                temporal_config=self.temporal_config,
            ).procesar()
            logs.extend(pse_result.get("logs", []))
            alertas.extend(pse_result.get("alertas", []))
        if contable_result is None and pse_result is None:
            raise ValueError("No se recibieron archivos para procesar")
        if contable_result is not None and pse_result is None:
            # Procesar Adquirencias si se proporciona (contable solo)
            if self.adquirencias_bytes is not None:
                try:
                    procesador_adq = ProcesadorAdquirencias(
                        self.adquirencias_bytes,
                        contable_result["file"],
                        value_tolerance=self.value_tolerance,
                    )
                    contable_result["file"] = procesador_adq.procesar()
                    logs.extend(procesador_adq.logs)
                except Exception as e:
                    alertas.append(f"Fallo en procesamiento de Adquirencias: {e}")
            return contable_result
        if contable_result is None and pse_result is not None:
            cruces_b64 = pse_result.get("secondary_file")
            if cruces_b64:
                cruces_b64 = self._annotate_pse_rows_in_contable(
                    cruces_b64,
                    pse_result.get("dataset", []),
                )
            files: list[dict[str, str]] = [
                {
                    "name": pse_result.get("secondary_output_name", "CRUCES_CONCILIADOS.xlsx"),
                    "file": cruces_b64,
                },
            ]
            return {
                "mode": "pse-only",
                "pse": pse_result,
                "files": files,
                "logs": logs,
                "alertas": alertas,
                "resumen": pse_result.get("resumen", {}),
            }
        contable_result["file"] = self._merge_pse_comments_into_contable(
            contable_result["file"],
            pse_result.get("dataset_cruces", []),
        )
        contable_result["file"] = self._annotate_pse_rows_in_contable(
            contable_result["file"],
            pse_result.get("dataset", []),
        )

        # Fase adicional: Conciliación por agrupación PSE (se ejecuta después de las reglas existentes)
        try:
            contable_result["file"] = self._apply_agrupacion_pse(
                contable_result["file"],
                pse_result.get("dataset", []),
            )
        except Exception as e:
            alertas.append(f"Fallo en agrupación PSE: {e}")

        # Fase adicional: Procesar Adquirencias si se proporciona
        if self.adquirencias_bytes is not None:
            try:
                procesador_adq = ProcesadorAdquirencias(
                    self.adquirencias_bytes,
                    contable_result["file"],
                    value_tolerance=self.value_tolerance,
                )
                contable_result["file"] = procesador_adq.procesar()
                logs.extend(procesador_adq.logs)
            except Exception as e:
                alertas.append(f"Fallo en procesamiento de Adquirencias: {e}")

        files: list[dict[str, str]] = [
            {
                "name": "CONCILIACION_CONTABLE.xlsx",
                "file": contable_result["file"],
            },
            {
                "name": pse_result.get("output_name", "PSE_CONCILIADO.xlsx"),
                "file": pse_result["file"],
            },
        ]
        resumen = {
            "cruzados": int(contable_result.get("resumen", {}).get("cruzados", 0)) + int(pse_result.get("resumen", {}).get("cruzados", 0)),
            "posibles": int(contable_result.get("resumen", {}).get("posibles", 0)) + int(pse_result.get("resumen", {}).get("posibles", 0)),
            "precision_estimada": round(
                (
                    (
                        int(contable_result.get("resumen", {}).get("cruzados", 0))
                        + int(pse_result.get("resumen", {}).get("cruzados", 0))
                    )
                    /
                    max(
                        1,
                        int(contable_result.get("resumen", {}).get("total_movements", 0))
                        + int(pse_result.get("resumen", {}).get("total_movements", 0)),
                    )
                ),
                2,
            ),
        }
        return {
            "mode": "integrado",
            "contable": contable_result,
            "pse": pse_result,
            "files": files,
            "logs": logs,
            "alertas": alertas,
            "resumen": resumen,
        }

    def _ensure_state_group_columns(self, workbook):
        """
        Asegura columnas: 'Estado_Conciliacion', 'ID_Grupo_Conciliacion' y devuelve mapping por sheet.
        """
        mapping: dict[str, dict[str, int]] = {}
        for sheet in workbook.worksheets:
            headers = {self._normalizar_texto(str(cell.value)): cell.column for cell in sheet[1] if cell.value}
            # buscar o crear Estado_Conciliacion
            estado_col = headers.get('estado_conciliacion')
            if estado_col is None:
                estado_col = sheet.max_column + 1
                sheet.cell(row=1, column=estado_col).value = 'Estado_Conciliacion'
            id_grp_col = headers.get('id_grupo_conciliacion')
            if id_grp_col is None:
                id_grp_col = sheet.max_column + 1
                sheet.cell(row=1, column=id_grp_col).value = 'ID_Grupo_Conciliacion'
            # comentario ya lo maneja _ensure_annotation_columns; buscar columna Comentario
            comment_col, observation_col = self._ensure_annotation_columns(workbook)[sheet.title]
            mapping[sheet.title] = {
                'valor': None,  # se calculará dinámicamente
                'estado': estado_col,
                'id_grupo': id_grp_col,
                'comentario': comment_col or observation_col,
                'descripcion': 2,
            }
        return mapping

    def _find_value_column(self, sheet) -> int:
        # buscar encabezado con palabras comunes
        candidates = ['valor','monto','importe','amount']
        for cell in sheet[1]:
            if isinstance(cell.value, str):
                h = self._normalizar_texto(cell.value)
                if any(c in h for c in candidates):
                    return cell.column
        # fallback: buscar primera columna con número en fila 2
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                if isinstance(cell.value, (int, float)):
                    return cell.column
        # si todo falla, usar columna 3
        return 3

    def _apply_agrupacion_pse(self, contable_b64: str, pse_dataset: list[dict[str, Any]]) -> str:
        workbook = load_workbook(filename=BytesIO(base64.b64decode(contable_b64)))
        # asegurar columnas y mapping
        sheet_maps = self._ensure_state_group_columns(workbook)
        annotation_cols_cache = self._ensure_annotation_columns(workbook)

        def _find_max_grp_number_in_workbook() -> int:
            maxn = 0
            for ws in workbook.worksheets:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and cell.startswith('GRP-'):
                            try:
                                num = int(cell.split('-')[1])
                                if num > maxn:
                                    maxn = num
                            except Exception:
                                continue
            return maxn

        max_grp_global = _find_max_grp_number_in_workbook()

        for sheet in workbook.worksheets:
            col_map = sheet_maps.get(sheet.title, {})
            # determinar columna de valor
            col_map['valor'] = self._find_value_column(sheet)
            # llamar a la función de agrupación por hoja
            conciliacion_por_agrupacion(
                sheet,
                pse_dataset,
                col_map,
                lambda: max_grp_global,
                lambda ws, r: self._row_contains_pse_marker(ws[r]),
            )
            max_grp_global = _find_max_grp_number_in_workbook()
        # Después de agrupar, anotar coincidencias en hojas de Bancolombia 1331 si aplica
        # Construir índice de grupos: {grp_id: [(sheet, row, value), ...]}
        groups: dict[str, list[tuple[str, int, float]]] = {}
        for sheet in workbook.worksheets:
            # determinar columnas para este sheet
            mapping = sheet_maps.get(sheet.title, {})
            id_col = mapping.get('id_grupo')
            val_col = mapping.get('valor')
            if not id_col or not val_col:
                continue
            for r in range(2, sheet.max_row + 1):
                gid = sheet.cell(row=r, column=id_col).value
                if not gid:
                    continue
                try:
                    v = float(sheet.cell(row=r, column=val_col).value or 0)
                except Exception:
                    v = 0.0
                groups.setdefault(str(gid), []).append((sheet.title, r, v))

        # heurística: buscar hojas relacionadas con Bancolombia/1331
        target_sheet_names = [name for name in workbook.sheetnames if self._is_bancolombia_sheet(name)]
        # anotar en las hojas objetivo si encuentran valores totales de grupo
        for gid, rows in groups.items():
            # seleccionar como total la fila con mayor valor absoluto
            if not rows:
                continue
            total_row = max(rows, key=lambda t: abs(t[2]))
            total_value = total_row[2]
            total_sheet, total_r, _ = total_row
            # preparar texto de comentario con detalle de desgloses
            breakdowns = [r for r in rows if not (r[0] == total_sheet and r[1] == total_r)]
            pse_count = len(breakdowns)
            breakdown_text = ", ".join(f"{s}:fila {row}={val:,.2f}" for s, row, val in breakdowns) if breakdowns else ""
            # mensaje similar al formato de ejemplo
            comment_text = (
                f"Cruce PSE ({gid}) - Conciliado N:1 contra {total_sheet} Cta 1331: fila {total_r}={total_value:,.2f}. "
                + (f"Pagos PSE asociados ({pse_count}): {breakdown_text}" if pse_count else "")
            )
            # buscar coincidencias en hojas objetivo
            for sheet_name in target_sheet_names:
                sheet = workbook[sheet_name]
                # intentar ubicar columna de valor en la hoja objetivo
                try:
                    target_val_col = self._find_value_column(sheet)
                except Exception:
                    target_val_col = 3
                for r in range(2, sheet.max_row + 1):
                    try:
                        v = float(sheet.cell(row=r, column=target_val_col).value or 0)
                    except Exception:
                        continue
                    if abs(abs(v) - abs(total_value)) <= self.value_tolerance:
                        # añadir comentario y observacion si existen columnas
                        comment_col, observation_col = annotation_cols_cache.get(sheet.title, (None, None))
                        if comment_col is not None:
                            ccell = sheet.cell(row=r, column=comment_col)
                            ccell.value = self._append_text_once(ccell.value, comment_text)
                        if observation_col is not None:
                            ocell = sheet.cell(row=r, column=observation_col)
                            ocell.value = self._append_text_once(ocell.value, f"({gid}) Pagos PSE asociados: {pse_count}")
        # Propagar grupos desde el dataset PSE a filas de Bancolombia/1331 (match por fecha+valor)
        pse_index: dict[tuple[str, int], set[str]] = {}
        pse_group_values: dict[str, list[float]] = {}
        for pse_row in pse_dataset:
            group_id = str(pse_row.get('id_grupo_conciliacion') or '').strip()
            if not group_id:
                continue
            pse_date = self._parse_date_text(str(pse_row.get('fecha') or ''))
            if pse_date is None:
                continue
            try:
                pse_value = abs(float(pse_row.get('valor') or pse_row.get('monto') or 0))
            except Exception:
                continue
            cents = int(round(pse_value * 100))
            pse_index.setdefault((pse_date.isoformat(), cents), set()).add(group_id)
            pse_group_values.setdefault(group_id, []).append(pse_value)

        for sheet_name in target_sheet_names:
            sheet = workbook[sheet_name]
            mapping = sheet_maps.get(sheet.title, {})
            id_col = mapping.get('id_grupo')
            estado_col = mapping.get('estado')
            if not id_col or not estado_col:
                continue
            comment_col, observation_col = annotation_cols_cache.get(sheet.title, (None, None))

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                row_idx = row[0].row
                estado_val = (sheet.cell(row=row_idx, column=estado_col).value or '').strip()
                if estado_val in ('Cruzado', 'Posible cruce'):
                    continue
                if sheet.cell(row=row_idx, column=id_col).value:
                    continue
                row_date_dt = self._parse_row_date(row)
                row_value = self._parse_row_value(row)
                if row_date_dt is None or row_value is None:
                    continue
                row_date = row_date_dt.date()
                cents = int(round(abs(row_value) * 100))

                best_candidate: tuple[int, int, str] | None = None
                for delta in range(-self.date_tolerance_days, self.date_tolerance_days + 1):
                    key_date_dt = row_date + timedelta(days=delta)
                    decision = evaluar_temporal(row_date, key_date_dt, self.temporal_config, None)
                    if not decision.permitida:
                        continue
                    gid_set = pse_index.get((key_date_dt.isoformat(), cents))
                    if not gid_set or len(gid_set) != 1:
                        continue
                    candidate_group = next(iter(gid_set))
                    candidate_rank = (decision.prioridad or 99, decision.diferencia_dias or 999, candidate_group)
                    if best_candidate is None or candidate_rank < best_candidate:
                        best_candidate = candidate_rank

                matched_group = best_candidate[2] if best_candidate is not None else None

                if not matched_group:
                    continue

                sheet.cell(row=row_idx, column=id_col).value = matched_group
                if not estado_val:
                    sheet.cell(row=row_idx, column=estado_col).value = 'Conciliado por agrupación'
                if comment_col is not None:
                    ccell = sheet.cell(row=row_idx, column=comment_col)
                    asociados = sorted(pse_group_values.get(matched_group, []), reverse=True)
                    asociados_text = ", ".join(f"{v:,.2f}" for v in asociados[:10])
                    n_asociados = len(asociados)
                    detalle_asociados = f"Pagos PSE asociados ({n_asociados}): {asociados_text}" if n_asociados else ""
                    comentario = (
                        f"Cruce PSE ({matched_group}) - Coincidencia por valor y fecha en {sheet.title}: fila {row_idx}={abs(row_value):,.2f}. "
                        + detalle_asociados
                    ).strip()
                    ccell.value = self._append_text_once(ccell.value, comentario)
                if observation_col is not None:
                    ocell = sheet.cell(row=row_idx, column=observation_col)
                    ocell.value = self._append_text_once(ocell.value, f"({matched_group})")

        out = BytesIO()
        workbook.save(out)
        return base64.b64encode(out.getvalue()).decode('utf-8')

