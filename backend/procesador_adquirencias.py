"""Procesador de Adquirencias: cruza valores, fechas y códigos con Bancolombia 690."""

from __future__ import annotations

import base64
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


LIGHT_BLUE_FILL = PatternFill(start_color="FFE8F4FF", end_color="FFE8F4FF", fill_type="solid")


class ProcesadorAdquirencias:
    def __init__(
        self,
        adquirencias_bytes: bytes,
        contable_b64: str,
        value_tolerance: float = 0.01,
        date_tolerance_days: int = 0,
    ) -> None:
        self.adquirencias_workbook = load_workbook(filename=BytesIO(adquirencias_bytes))
        self.contable_workbook = load_workbook(filename=BytesIO(base64.b64decode(contable_b64)))
        self.value_tolerance = value_tolerance
        self.date_tolerance_days = max(0, int(date_tolerance_days))
        self.logs: list[dict[str, Any]] = []
        self.adquirencia_counter = 1
        self._ensure_annotation_columns(self.adquirencias_workbook)
        self._ensure_annotation_columns(self.contable_workbook)
        self.annotation_columns_adq = self._index_annotation_columns(self.adquirencias_workbook)
        self.annotation_columns_cont = self._index_annotation_columns(self.contable_workbook)

    def procesar(self) -> dict[str, str]:
        """Procesa Adquirencias y retorna dict con archivos en base64."""
        adquirencias_data = self._extraer_adquirencias_con_fila()
        self._cruzar_ambos_archivos(adquirencias_data)
        
        adq_stream = BytesIO()
        self.adquirencias_workbook.save(adq_stream)
        adq_b64 = base64.b64encode(adq_stream.getvalue()).decode("utf-8")
        
        cont_stream = BytesIO()
        self.contable_workbook.save(cont_stream)
        cont_b64 = base64.b64encode(cont_stream.getvalue()).decode("utf-8")
        
        return {
            "adquirencias_file": adq_b64,
            "contable_file": cont_b64,
        }

    def _normalizar_texto(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", str(value) if value else "")
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _index_annotation_columns(self, workbook) -> dict[str, tuple[int | None, int | None]]:
        indexed: dict[str, tuple[int | None, int | None]] = {}
        for sheet in workbook.worksheets:
            comment_col: int | None = None
            observation_col: int | None = None
            max_scan_row = min(sheet.max_row, 25)
            for row in sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)
                    if comment_col is None and "coment" in header:
                        comment_col = cell.column
                    if observation_col is None and "observ" in header:
                        observation_col = cell.column
                if comment_col is not None and observation_col is not None:
                    break
            indexed[sheet.title] = (comment_col, observation_col)
        return indexed

    def _ensure_annotation_columns(self, workbook) -> None:
        for sheet in workbook.worksheets:
            headers = {
                self._normalizar_texto(str(cell.value)): cell.column
                for cell in sheet[1]
                if cell.value is not None
            }

            if "coment" not in " ".join(headers.keys()):
                next_col = sheet.max_column + 1
                sheet.cell(row=1, column=next_col).value = "Comentario"

            headers = {
                self._normalizar_texto(str(cell.value)): cell.column
                for cell in sheet[1]
                if cell.value is not None
            }

            if "observ" not in " ".join(headers.keys()):
                next_col = sheet.max_column + 1
                sheet.cell(row=1, column=next_col).value = "Observacion"

    def _parse_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(cleaned, fmt).date()
                    except Exception:
                        continue
        return None

    def _parse_amount(self, value: Any) -> float | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if not isinstance(value, str):
            return None

        raw = value.strip()
        if not raw:
            return None

        cleaned = raw.replace(" ", "").replace("$", "")
        if "," in cleaned and "." in cleaned:
            # Soporta formatos 1.234,56 y 1,234.56
            if cleaned.rfind(",") > cleaned.rfind("."):
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")

        try:
            return float(cleaned)
        except Exception:
            return None

    def _normalizar_autorizacion(self, value: Any) -> str:
        if value is None:
            return ""
        normalized = self._normalizar_texto(str(value))
        return "".join(ch for ch in normalized if ch.isalnum())

    def _extraer_token_aprobacion(self, value: Any) -> str:
        if value is None:
            return ""

        text = self._normalizar_texto(str(value))
        if not text:
            return ""

        tokens = re.findall(r"[a-z0-9]{4,30}", text)
        if not tokens:
            return ""

        def token_score(token: str) -> tuple[int, int, int]:
            digits = sum(ch.isdigit() for ch in token)
            has_digits = 1 if digits > 0 else 0
            # Los codigos de aprobacion suelen ser compactos; penaliza cuentas largas.
            compact_bonus = 1 if 4 <= len(token) <= 8 else 0
            long_penalty = -1 if len(token) > 10 and digits == len(token) else 0
            return (has_digits + compact_bonus + long_penalty, digits, -len(token))

        tokens_with_digits = [token for token in tokens if any(ch.isdigit() for ch in token)]
        ranked = tokens_with_digits if tokens_with_digits else tokens
        ranked.sort(key=token_score, reverse=True)
        return ranked[0]

    def _extraer_aprobacion_en_fila(self, row, preferred_col: int | None = None) -> str:
        if preferred_col is not None and preferred_col > 0 and preferred_col - 1 < len(row):
            direct = self._extraer_token_aprobacion(row[preferred_col - 1].value)
            if direct:
                return direct

        candidates: list[str] = []
        for cell in row:
            token = self._extraer_token_aprobacion(cell.value)
            if token:
                candidates.append(token)

        if not candidates:
            return ""

        candidates.sort(key=lambda token: (sum(ch.isdigit() for ch in token), 1 if 4 <= len(token) <= 8 else 0, -len(token)), reverse=True)
        return candidates[0]

    def _parse_date_from_row(self, row) -> date | None:
        for cell in row:
            parsed = self._parse_date(cell.value)
            if parsed is not None:
                return parsed
        return None

    def _parse_value_candidates(self, row, exclude_cols: set[int] | None = None) -> list[tuple[int, float, Any]]:
        candidates: list[tuple[int, float, Any]] = []
        excluded = exclude_cols or set()
        for idx, cell in enumerate(row, start=1):
            if idx in excluded:
                continue
            value = cell.value
            if value is None or isinstance(value, bool):
                continue
            if getattr(cell, "is_date", False):
                continue
            parsed = self._parse_amount(value)
            if parsed is None or abs(parsed) < 0.0001:
                continue
            candidates.append((idx, float(parsed), cell))
        return candidates

    def _extraer_movimientos_690(self) -> tuple[Worksheet | None, list[dict[str, Any]]]:
        sheet = None
        for candidate in self.contable_workbook.worksheets:
            title = self._normalizar_texto(candidate.title)
            if "690" in title or "1331" in title or "bancolombia" in title:
                sheet = candidate
                break

        if sheet is None:
            return None, []

        marker_candidates = [
            "consignaciones sin registrar",
            "sin registrar en el extracto",
            "sin registrar en libros",
        ]
        start_row = None
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            texts = [self._normalizar_texto(str(cell.value)) for cell in row if isinstance(cell.value, str) and str(cell.value).strip()]
            if any(any(marker in text for marker in marker_candidates) for text in texts):
                start_row = row[0].row + 1
                break

        if start_row is None:
            start_row = 2

        # Para 690, usar estructura estándar de Bancolombia:
        # Columna B (idx 1): Fecha
        # Columna D (idx 3): Valor de consignación
        # Columna F (idx 5): Código de aprobación
        # No hacer detección automática, usar posiciones fijas.
        
        valor_col = 4  # Columna D
        fecha_col = 2  # Columna B
        
        entries: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
            try:
                # Leer columna B (índice 1) para fecha
                fecha_cell = row[1] if len(row) > 1 else None
                raw_date = self._parse_date(fecha_cell.value) if fecha_cell else None
                if raw_date is None:
                    continue
                
                # Leer columna D (índice 3) para valor
                valor_cell = row[3] if len(row) > 3 else None
                if valor_cell is None:
                    continue
                value = self._parse_amount(valor_cell.value)
                if value is None or abs(value) < 0.0001:
                    continue
                
                # Leer código de aprobación SIEMPRE de columna F (índice 5)
                approval = ""
                if len(row) > 5:
                    approval = self._normalizar_autorizacion(row[5].value)
                if not approval:
                    approval = self._extraer_aprobacion_en_fila(row)
                if not approval:
                    continue

                entries.append(
                    {
                        "sheet": sheet.title,
                        "row": row[0].row,
                        "valor": float(value),
                        "fecha": raw_date,
                        "autorizacion": approval,
                        "valor_cell": valor_cell,
                        "valor_col": valor_col,
                    }
                )
            except Exception:
                continue

        return sheet, entries

    def _find_header_columns(
        self,
        sheet: Worksheet,
        *,
        start_row: int = 1,
        need_auth: bool = True,
    ) -> tuple[int | None, int | None, int | None, int]:
        valor_col: int | None = None
        fecha_col: int | None = None
        auth_col: int | None = None
        header_row = 1

        valor_preferidos = ["valor total", "valor neto"]
        fecha_preferidas = ["fecha de transaccion", "fecha de compensacion"]
        auth_preferidas = ["codigo autorizacion", "codigo de autorizacion", "numero aprobacion", "nro aprobacion"]

        valor_candidatos: list[tuple[int, int]] = []
        fecha_candidatos: list[tuple[int, int]] = []
        auth_candidatos: list[tuple[int, int]] = []

        def score_header(header: str, preferred: list[str]) -> int:
            score = 0
            for index, token in enumerate(preferred):
                if token == header:
                    return 1000 - index
                if token in header:
                    score = max(score, 500 - index)
            return score

        max_scan = min(sheet.max_row, start_row + 29)
        for row_idx in range(start_row, max_scan + 1):
            row = sheet[row_idx]
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                header = self._normalizar_texto(cell.value)
                if any(x in header for x in {"valor", "monto", "importe", "amount", "consignacion"}):
                    valor_candidatos.append((score_header(header, valor_preferidos), cell.column))
                if any(x in header for x in {"fecha", "date", "transaccion", "movimiento", "transaction"}):
                    fecha_candidatos.append((score_header(header, fecha_preferidas), cell.column))
                if any(
                    x in header
                    for x in {
                        "autorizacion",
                        "authorization",
                        "codigo autoriz",
                        "cod autoriz",
                        "auth",
                        "reference",
                        "aprobacion",
                        "numero aprob",
                        "num aprob",
                        "nro aprob",
                        "num operacion",
                        "numero operacion",
                        "id transaccion",
                        "id trx",
                        "referencia",
                        "trace",
                        "nsu",
                    }
                ):
                    auth_candidatos.append((score_header(header, auth_preferidas), cell.column))
            if valor_col is None and valor_candidatos:
                valor_col = max(valor_candidatos, key=lambda item: (item[0], -item[1]))[1]
            if fecha_col is None and fecha_candidatos:
                fecha_col = max(fecha_candidatos, key=lambda item: (item[0], -item[1]))[1]
            if auth_col is None and auth_candidatos:
                auth_col = max(auth_candidatos, key=lambda item: (item[0], -item[1]))[1]
            if valor_col is not None and fecha_col is not None and (not need_auth or auth_col is not None):
                header_row = row_idx
                break

        return valor_col, fecha_col, auth_col, header_row

    def _find_header_columns_from_row(
        self,
        sheet: Worksheet,
        *,
        start_row: int = 1,
        need_auth: bool = True,
    ) -> tuple[int | None, int | None, int | None, int]:
        """Alias para _find_header_columns con start_row especificado."""
        return self._find_header_columns(sheet, start_row=start_row, need_auth=need_auth)

    def _extraer_adquirencias_con_fila(self) -> list[dict[str, Any]]:
        """Extrae datos de Adquirencias con referencias a celdas."""
        datos = []
        for sheet in self.adquirencias_workbook.worksheets:
            valor_col, fecha_col, auth_col, header_row = self._find_header_columns(sheet, need_auth=False)
            
            if valor_col is None or fecha_col is None:
                continue
            
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
                try:
                    valor_cell = row[valor_col - 1] if valor_col - 1 < len(row) else None
                    fecha_cell = row[fecha_col - 1] if fecha_col - 1 < len(row) else None
                    if valor_cell is None or fecha_cell is None:
                        continue
                    
                    valor = self._parse_amount(valor_cell.value)
                    if valor is None:
                        continue
                    
                    if abs(valor) < 0.0001:
                        continue
                    
                    fecha = self._parse_date(fecha_cell.value)
                    if fecha is None:
                        continue
                    
                    auth_code = ""
                    if auth_col is not None and auth_col > 0 and auth_col - 1 < len(row):
                        auth_code = self._normalizar_autorizacion(row[auth_col - 1].value)
                    if not auth_code:
                        auth_code = self._extraer_aprobacion_en_fila(row, preferred_col=auth_col)
                    if not auth_code:
                        continue
                    
                    datos.append({
                        "sheet": sheet.title,
                        "row": row[0].row,
                        "valor": valor,
                        "fecha": fecha,
                        "autorizacion": auth_code,
                        "valor_cell": valor_cell,
                        "valor_col": valor_col,
                    })
                except Exception:
                    continue
        
        return datos

    def _cruzar_ambos_archivos(self, adquirencias_data: list[dict[str, Any]]) -> None:
        """Busca coincidencias por número de aprobación y valida fecha+valor exactos."""
        bancolombia_sheet, bancolombia_entries = self._extraer_movimientos_690()

        if bancolombia_sheet is None:
            return
        
        adq_by_auth: dict[str, list[dict]] = {}
        for adq in adquirencias_data:
            adq_by_auth.setdefault(adq["autorizacion"], []).append(adq)
        
        matched_adq_rows: set[tuple[str, int]] = set()
        for bank_entry in bancolombia_entries:
            try:
                banco_valor = bank_entry["valor"]
                banco_fecha = bank_entry["fecha"]
                banco_auth = bank_entry["autorizacion"]

                if banco_fecha is None or not banco_auth:
                    continue
                
                matching_adq = adq_by_auth.get(banco_auth, [])
                if not matching_adq:
                    continue

                adq = None
                for candidate in matching_adq:
                    adq_row_key = (candidate["sheet"], candidate["row"])
                    if adq_row_key in matched_adq_rows:
                        continue
                    day_delta = abs((candidate["fecha"] - banco_fecha).days)
                    if day_delta > self.date_tolerance_days:
                        continue
                    value_delta = abs(round(abs(candidate["valor"]), 2) - round(abs(banco_valor), 2))
                    if value_delta > self.value_tolerance:
                        continue
                    adq = candidate
                    break

                if adq is None:
                    continue

                matched_adq_rows.add((adq["sheet"], adq["row"]))
                
                tag = f"Adquirencia {self.adquirencia_counter}"
                self.adquirencia_counter += 1
                
                bank_entry["valor_cell"].fill = LIGHT_BLUE_FILL
                adq["valor_cell"].fill = LIGHT_BLUE_FILL
                
                self._anotar_cruce(
                    bancolombia_sheet,
                    bank_entry["row"],
                    tag,
                    self.annotation_columns_cont,
                    f"Encontrado en {adq['sheet']}:fila {adq['row']}"
                )
                
                adq_sheet = self.adquirencias_workbook[adq["sheet"]]
                self._anotar_cruce(
                    adq_sheet,
                    adq["row"],
                    tag,
                    self.annotation_columns_adq,
                    f"Encontrado en {bancolombia_sheet.title}:fila {bank_entry['row']}"
                )
                
                self.logs.append({
                    "tipo": "adquirencia_cruzada",
                    "valor": round(abs(banco_valor), 2),
                    "fecha": banco_fecha.isoformat(),
                    "confianza": 0.95,
                    "detalle": (
                        f"{tag}: coincidencia por aprobacion {banco_auth}, "
                        f"valor {banco_valor:,.2f} y fecha {banco_fecha.isoformat()} "
                        f"en {bancolombia_sheet.title}:fila {bank_entry['row']}"
                    ),
                })
                
            except Exception:
                continue

    def _anotar_cruce(
        self,
        sheet: Worksheet,
        row_idx: int,
        tag: str,
        annotation_cols: dict[str, tuple[int | None, int | None]],
        location_text: str,
    ) -> None:
        comment_col, observation_col = annotation_cols.get(sheet.title, (None, None))
        
        if comment_col is not None:
            comment_cell = sheet.cell(row=row_idx, column=comment_col)
            comment_text = f"{tag} - Cruce con Adquirencias"
            current = str(comment_cell.value).strip() if comment_cell.value else ""
            if current and tag not in current:
                comment_cell.value = f"{current} | {comment_text}"
            elif not current:
                comment_cell.value = comment_text
        
        if observation_col is not None:
            observation_cell = sheet.cell(row=row_idx, column=observation_col)
            observation_text = f"{tag} {location_text}"
            current = str(observation_cell.value).strip() if observation_cell.value else ""
            if current and tag not in current:
                observation_cell.value = f"{current} | {observation_text}"
            elif not current:
                observation_cell.value = observation_text
