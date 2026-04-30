from __future__ import annotations

import base64
import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet


DATE_ALIASES = {
    "fecha",
    "date",
    "transaction date",
    "fecha movimiento",
    "fecha transaccion",
    "fecha transacción",
}
VALUE_ALIASES = {
    "valor",
    "monto",
    "importe",
    "amount",
    "valor transaccion",
    "valor transacción",
    "valor movimiento",
}
ACCOUNT_ALIASES = {
    "cuenta",
    "cuenta contable",
    "account",
    "codigo cuenta",
    "código cuenta",
    "numero cuenta",
    "número cuenta",
}

VALUE_TOLERANCE_DEFAULT = 0.01
DATE_TOLERANCE_DEFAULT = 1
MAX_SUBSET_CANDIDATES = 16
MAX_SUBSET_SIZE = 6
PARTIAL_THRESHOLD_RATIO = 0.015

HEADER_FILL = PatternFill(start_color="FFEAF2FF", end_color="FFEAF2FF", fill_type="solid")


@dataclass
class SheetSchema:
    header_row: int
    date_col: int
    value_col: int
    account_col: int | None = None


@dataclass
class Movimiento:
    sheet_name: str
    row: int
    raw_date: date | None
    value: float
    account_label: str
    matched: bool = False


@dataclass
class MatchResult:
    pse_entry: Movimiento
    matched_entries: list[Movimiento]
    state: str
    group_id: str | None
    comment: str
    account_label: str
    values_associated: str
    difference: float | None


class PseConciliador:
    def __init__(
        self,
        pse_bytes: bytes,
        cruces_bytes: bytes,
        *,
        date_tolerance_days: int = DATE_TOLERANCE_DEFAULT,
        value_tolerance: float = VALUE_TOLERANCE_DEFAULT,
    ) -> None:
        self.pse_workbook = load_workbook(filename=BytesIO(pse_bytes))
        self.cruces_workbook = load_workbook(filename=BytesIO(cruces_bytes))
        self.date_tolerance_days = max(0, int(date_tolerance_days))
        self.value_tolerance = max(0.0, float(value_tolerance))
        self.logs: list[dict[str, Any]] = []
        self.alertas: list[str] = []
        self.dataset: list[dict[str, Any]] = []
        self.dataset_cruces: list[dict[str, Any]] = []
        self.group_counter = 1
        self.pse_schemas: dict[str, SheetSchema] = {}
        self.cruces_schemas: dict[str, SheetSchema] = {}
        self.pse_entries: list[Movimiento] = []
        self.cruces_entries: list[Movimiento] = []
        self.match_results: list[MatchResult] = []

    def procesar(self) -> dict[str, Any]:
        self.pse_entries = self._extract_movements(self.pse_workbook, self.pse_schemas)
        self.cruces_entries = self._extract_movements(self.cruces_workbook, self.cruces_schemas)
        self.match_results = self._conciliar(self.pse_entries, self.cruces_entries)

        self._escribir_enriquecimiento_pse(self.match_results)
        self._escribir_enriquecimiento_cruces(self.match_results)
        resumen = self._construir_resumen(self.match_results, self.pse_entries)
        self.alertas = self._construir_alertas(self.match_results, resumen)

        pse_output = BytesIO()
        self.pse_workbook.save(pse_output)
        pse_b64 = base64.b64encode(pse_output.getvalue()).decode("utf-8")

        cruces_output = BytesIO()
        self.cruces_workbook.save(cruces_output)
        cruces_b64 = base64.b64encode(cruces_output.getvalue()).decode("utf-8")

        return {
            "file": pse_b64,
            "output_name": "PSE_CONCILIADO.xlsx",
            "secondary_file": cruces_b64,
            "secondary_output_name": "CRUCES_CONCILIADOS.xlsx",
            "dataset": self.dataset,
            "dataset_cruces": self.dataset_cruces,
            "logs": self.logs,
            "resumen": resumen,
            "alertas": self.alertas,
        }

    def _normalizar_texto(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value)
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _parse_date(self, cell) -> date | None:
        value = cell.value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if getattr(cell, "is_date", False):
            if isinstance(value, (int, float)):
                try:
                    parsed = from_excel(value)
                    return parsed.date() if isinstance(parsed, datetime) else parsed
                except Exception:
                    return None
        return None

    def _parse_float(self, value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace("$", "").replace(" ", "")
            if not cleaned:
                return None
            last_comma = cleaned.rfind(",")
            last_dot = cleaned.rfind(".")

            if last_comma != -1 and last_dot != -1:
                if last_comma > last_dot:
                    cleaned = cleaned.replace(".", "").replace(",", ".")
                else:
                    cleaned = cleaned.replace(",", "")
            elif cleaned.count(",") == 1:
                decimals = len(cleaned.split(",")[-1])
                cleaned = cleaned.replace(",", ".") if decimals <= 2 else cleaned.replace(",", "")
            elif cleaned.count(".") == 1:
                decimals = len(cleaned.split(".")[-1])
                cleaned = cleaned if decimals <= 2 else cleaned.replace(".", "")
            elif cleaned.count(".") > 1:
                cleaned = cleaned.replace(".", "")
            else:
                cleaned = cleaned.replace(",", "")
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _sheet_account_label(self, sheet_name: str) -> str:
        match = re.search(r"\d{4,}", sheet_name)
        if match:
            return match.group(0)
        return sheet_name.strip()

    def _detect_schema(self, sheet: Worksheet) -> SheetSchema | None:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), min_col=1, max_col=sheet.max_column):
            header_map: dict[str, int] = {}
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                header = self._normalizar_texto(cell.value)
                if not header:
                    continue
                if (header in DATE_ALIASES or any(alias in header for alias in DATE_ALIASES)) and "date" not in header_map:
                    header_map["date"] = cell.column
                if (header in VALUE_ALIASES or any(alias in header for alias in VALUE_ALIASES)) and "value" not in header_map:
                    header_map["value"] = cell.column
                if (header in ACCOUNT_ALIASES or any(alias in header for alias in ACCOUNT_ALIASES)) and "account" not in header_map:
                    header_map["account"] = cell.column

            if "date" in header_map and "value" in header_map:
                return SheetSchema(
                    header_row=row[0].row,
                    date_col=header_map["date"],
                    value_col=header_map["value"],
                    account_col=header_map.get("account"),
                )
        return None

    def _resolve_account_label(self, sheet: Worksheet, row, schema: SheetSchema) -> str:
        if schema.account_col is not None:
            cell_value = row[schema.account_col - 1].value if schema.account_col - 1 < len(row) else None
            cell_text = self._stringify(cell_value)
            if cell_text:
                return cell_text

        for cell in row:
            text = self._stringify(cell.value)
            if text and re.search(r"\d{4,}", text):
                return text

        return self._sheet_account_label(sheet.title)

    def _extract_movements(
        self,
        workbook,
        schema_registry: dict[str, SheetSchema],
    ) -> list[Movimiento]:
        movements: list[Movimiento] = []

        for sheet in workbook.worksheets:
            schema = self._detect_schema(sheet)
            if schema is None:
                continue

            schema_registry[sheet.title] = schema

            for row in sheet.iter_rows(min_row=schema.header_row + 1, max_row=sheet.max_row):
                date_cell = row[schema.date_col - 1] if schema.date_col - 1 < len(row) else None
                value_cell = row[schema.value_col - 1] if schema.value_col - 1 < len(row) else None
                if date_cell is None or value_cell is None:
                    continue

                raw_date = self._parse_date(date_cell)
                value = self._parse_float(value_cell.value)
                if raw_date is None or value is None or abs(value) < 0.0001:
                    continue

                movements.append(
                    Movimiento(
                        sheet_name=sheet.title,
                        row=row[0].row,
                        raw_date=raw_date,
                        value=value,
                        account_label=self._resolve_account_label(sheet, row, schema),
                    )
                )

        return movements

    def _date_gap_days(self, left: date | None, right: date | None) -> int | None:
        if left is None or right is None:
            return None
        return abs((left - right).days)

    def _within_value_tolerance(self, left: float, right: float) -> bool:
        return abs(abs(left) - abs(right)) <= self.value_tolerance

    def _currency_string(self, value: float) -> str:
        return f"{abs(value):,.2f}"

    def _candidate_pool(self, entry: Movimiento, pool: list[Movimiento]) -> list[Movimiento]:
        candidates: list[Movimiento] = []
        for candidate in pool:
            if candidate.matched:
                continue
            if candidate.raw_date is None:
                continue
            gap = self._date_gap_days(entry.raw_date, candidate.raw_date)
            if gap is None or gap > self.date_tolerance_days:
                continue
            if abs(candidate.value) > abs(entry.value) + self.value_tolerance:
                continue
            candidates.append(candidate)

        candidates.sort(
            key=lambda item: (
                self._date_gap_days(entry.raw_date, item.raw_date) or 999,
                abs(abs(entry.value) - abs(item.value)),
                -abs(item.value),
            )
        )
        return candidates[:MAX_SUBSET_CANDIDATES]

    def _find_exact_match(self, entry: Movimiento, pool: list[Movimiento]) -> Movimiento | None:
        candidates = [
            candidate
            for candidate in pool
            if not candidate.matched
            and candidate.raw_date is not None
            and self._date_gap_days(entry.raw_date, candidate.raw_date) is not None
            and self._date_gap_days(entry.raw_date, candidate.raw_date) <= self.date_tolerance_days
            and self._within_value_tolerance(entry.value, candidate.value)
        ]
        candidates.sort(
            key=lambda item: (
                self._date_gap_days(entry.raw_date, item.raw_date) or 999,
                abs(abs(entry.value) - abs(item.value)),
                item.row,
            )
        )
        return candidates[0] if candidates else None

    def _find_best_subset(
        self,
        entry: Movimiento,
        candidates: list[Movimiento],
    ) -> tuple[list[Movimiento] | None, float | None, bool]:
        if not candidates:
            return None, None, False

        target_cents = int(round(abs(entry.value) * 100))
        tolerance_cents = max(1, int(round(self.value_tolerance * 100)))
        best_exact: list[Movimiento] | None = None
        best_partial: list[Movimiento] | None = None
        best_partial_diff: int | None = None
        found_exact = False

        def backtrack(start: int, chosen: list[Movimiento], current_sum: int) -> None:
            nonlocal best_exact, best_partial, best_partial_diff, found_exact

            if found_exact:
                return

            if chosen:
                diff = abs(target_cents - current_sum)
                if diff <= tolerance_cents:
                    best_exact = chosen.copy()
                    found_exact = True
                    return
                if best_partial_diff is None or diff < best_partial_diff:
                    best_partial_diff = diff
                    best_partial = chosen.copy()

            if len(chosen) >= MAX_SUBSET_SIZE or start >= len(candidates):
                return

            for index in range(start, len(candidates)):
                candidate = candidates[index]
                chosen.append(candidate)
                backtrack(index + 1, chosen, current_sum + int(round(abs(candidate.value) * 100)))
                chosen.pop()
                if found_exact:
                    return

        backtrack(0, [], 0)

        if best_exact is not None:
            return best_exact, 0.0, True

        if best_partial is None or best_partial_diff is None:
            return None, None, False

        partial_threshold = max(tolerance_cents * 10, int(round(target_cents * PARTIAL_THRESHOLD_RATIO)))
        if best_partial_diff <= partial_threshold:
            return best_partial, best_partial_diff / 100.0, False

        return None, None, False

    def _build_account_string(self, entries: list[Movimiento]) -> str:
        labels = []
        for entry in entries:
            if entry.account_label and entry.account_label not in labels:
                labels.append(entry.account_label)
        return ", ".join(labels)

    def _build_associated_values(self, entries: list[Movimiento]) -> str:
        parts = []
        for entry in entries:
            parts.append(f"{entry.sheet_name}:fila {entry.row}={self._currency_string(entry.value)}")
        return " | ".join(parts)

    def _build_result_map(self, match_results: list[MatchResult]) -> dict[tuple[str, int], MatchResult]:
        result_map: dict[tuple[str, int], MatchResult] = {}
        for result in match_results:
            result_map[(result.pse_entry.sheet_name, result.pse_entry.row)] = result
            for entry in result.matched_entries:
                result_map[(entry.sheet_name, entry.row)] = result
        return result_map

    def _prepare_output_sheet(self, sheet: Worksheet, header_row: int, headers: list[str]) -> int:
        start_column = sheet.max_column + 1
        for offset, header in enumerate(headers):
            target_column = start_column + offset
            header_cell = sheet.cell(row=header_row, column=target_column)
            header_cell.value = header
            self._style_new_header(sheet, header_row, target_column)
        return start_column

    def _write_enrichment_rows(
        self,
        sheet: Worksheet,
        schema: SheetSchema,
        movements: list[Movimiento],
        result_map: dict[tuple[str, int], MatchResult],
        *,
        dataset_target: list[dict[str, Any]],
    ) -> None:
        start_column = self._prepare_output_sheet(
            sheet,
            schema.header_row,
            [
                "Estado_Conciliacion",
                "Cuenta_Contable",
                "Valores_Asociados",
                "Comentario_Conciliacion",
                "ID_Grupo_Conciliacion",
            ],
        )

        for movement in movements:
            if movement.sheet_name != sheet.title:
                continue

            result = result_map.get((movement.sheet_name, movement.row))
            if result is None:
                state = "Sin coincidencia"
                account_label = ""
                values_associated = ""
                comment = "Sin coincidencia en cruces contables"
                group_id = ""
                difference = None
            else:
                state = result.state
                account_label = result.account_label
                values_associated = result.values_associated
                comment = result.comment
                group_id = result.group_id or ""
                difference = result.difference

            sheet.cell(row=movement.row, column=start_column).value = state
            sheet.cell(row=movement.row, column=start_column + 1).value = account_label
            sheet.cell(row=movement.row, column=start_column + 2).value = values_associated
            sheet.cell(row=movement.row, column=start_column + 3).value = comment
            sheet.cell(row=movement.row, column=start_column + 4).value = group_id

            dataset_target.append(
                {
                    "sheet": movement.sheet_name,
                    "row": movement.row,
                    "estado_conciliacion": state,
                    "cuenta_contable": account_label,
                    "valores_asociados": values_associated,
                    "comentario_conciliacion": comment,
                    "id_grupo_conciliacion": group_id,
                    "diferencia": round(difference, 2) if difference is not None else None,
                }
            )

    def _escribir_enriquecimiento_pse(self, match_results: list[MatchResult]) -> None:
        result_map = self._build_result_map(match_results)
        for sheet in self.pse_workbook.worksheets:
            schema = self.pse_schemas.get(sheet.title)
            if schema is None:
                continue
            self._write_enrichment_rows(
                sheet,
                schema,
                self.pse_entries,
                result_map,
                dataset_target=self.dataset,
            )

    def _escribir_enriquecimiento_cruces(self, match_results: list[MatchResult]) -> None:
        result_map = self._build_result_map(match_results)
        for sheet in self.cruces_workbook.worksheets:
            schema = self.cruces_schemas.get(sheet.title)
            if schema is None:
                continue
            self._write_enrichment_rows(
                sheet,
                schema,
                self.cruces_entries,
                result_map,
                dataset_target=self.dataset_cruces,
            )

    def _register_match(
        self,
        pse_entry: Movimiento,
        matched_entries: list[Movimiento],
        state: str,
        difference: float | None,
    ) -> MatchResult:
        group_id = f"GRP-{self.group_counter:04d}"
        self.group_counter += 1

        for entry in matched_entries:
            entry.matched = True
        pse_entry.matched = True

        account_label = self._build_account_string(matched_entries)
        values_associated = self._build_associated_values(matched_entries)
        total_match_value = sum(abs(entry.value) for entry in matched_entries)

        if state == "Conciliado 1:1":
            comment = (
                f"Conciliado con cuenta {account_label} por valor total {self._currency_string(total_match_value)} "
                f"(1 transacción contable). Pertenece a PSE {pse_entry.sheet_name}:fila {pse_entry.row} "
                f"valor {self._currency_string(pse_entry.value)}"
            )
            log_type = "match_1_1"
        elif state == "Conciliado 1:N":
            comment = (
                f"Conciliado con cuenta {account_label} por valor total {self._currency_string(total_match_value)} "
                f"({len(matched_entries)} transacciones contables). Pertenece a PSE {pse_entry.sheet_name}:fila {pse_entry.row} "
                f"valor {self._currency_string(pse_entry.value)}"
            )
            log_type = "match_1_n"
        else:
            comment = (
                f"Conciliación parcial - diferencia de {self._currency_string(difference or 0.0)}. "
                f"Pertenece a PSE {pse_entry.sheet_name}:fila {pse_entry.row} valor {self._currency_string(pse_entry.value)}"
            )
            log_type = "partial"

        self.logs.append(
            {
                "tipo": log_type,
                "valor": round(abs(pse_entry.value), 2),
                "fecha": pse_entry.raw_date.isoformat() if pse_entry.raw_date else None,
                "confianza": 0.99 if state.startswith("Conciliado") else 0.7,
                "detalle": f"{state} entre {pse_entry.sheet_name}:fila {pse_entry.row} y {values_associated}",
            }
        )

        return MatchResult(
            pse_entry=pse_entry,
            matched_entries=matched_entries,
            state=state,
            group_id=group_id,
            comment=comment,
            account_label=account_label,
            values_associated=values_associated,
            difference=difference,
        )

    def _conciliar(self, pse_entries: list[Movimiento], cruce_entries: list[Movimiento]) -> list[MatchResult]:
        results: list[MatchResult] = []
        ordered_entries = sorted(pse_entries, key=lambda item: abs(item.value), reverse=True)

        for entry in ordered_entries:
            if entry.matched:
                continue

            exact_match = self._find_exact_match(entry, cruce_entries)
            if exact_match is not None:
                results.append(self._register_match(entry, [exact_match], "Conciliado 1:1", None))
                continue

            candidates = self._candidate_pool(entry, cruce_entries)
            subset, difference, is_exact = self._find_best_subset(entry, candidates)
            if subset is not None:
                state = "Conciliado 1:N" if is_exact and len(subset) > 1 else "Conciliación parcial"
                results.append(self._register_match(entry, subset, state, difference))
                continue

            results.append(
                MatchResult(
                    pse_entry=entry,
                    matched_entries=[],
                    state="Sin coincidencia",
                    group_id=None,
                    comment="Sin coincidencia en cruces contables",
                    account_label="",
                    values_associated="",
                    difference=None,
                )
            )
            self.logs.append(
                {
                    "tipo": "unmatched",
                    "valor": round(abs(entry.value), 2),
                    "fecha": entry.raw_date.isoformat() if entry.raw_date else None,
                    "confianza": 0.0,
                    "detalle": f"Sin coincidencia en cruces contables para {entry.sheet_name}:fila {entry.row}",
                }
            )

        return results

    def _style_new_header(self, sheet: Worksheet, header_row: int, column_index: int) -> None:
        source_column = max(1, column_index - 1)
        source_cell = sheet.cell(row=header_row, column=source_column)
        target_cell = sheet.cell(row=header_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = HEADER_FILL
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    def _construir_resumen(self, match_results: list[MatchResult], pse_entries: list[Movimiento]) -> dict[str, Any]:
        conciliados = len([result for result in match_results if result.state.startswith("Conciliado")])
        parciales = len([result for result in match_results if result.state == "Conciliación parcial"])
        sin_coincidencia = len([result for result in match_results if result.state == "Sin coincidencia"])
        total = len(pse_entries)
        porcentaje_conciliados = round((conciliados / total * 100) if total else 0.0, 2)
        porcentaje_parciales = round((parciales / total * 100) if total else 0.0, 2)
        precision = round((conciliados / total) if total else 0.0, 2)

        return {
            "cruzados": conciliados,
            "posibles": parciales,
            "precision_estimada": precision,
            "total_movements": total,
            "porcentaje_cruzados": porcentaje_conciliados,
            "porcentaje_posibles": porcentaje_parciales,
            "sin_coincidencia": sin_coincidencia,
        }

    def _construir_alertas(self, match_results: list[MatchResult], resumen: dict[str, Any]) -> list[str]:
        alertas: list[str] = []
        conciliados = int(resumen["cruzados"])
        sin_coincidencia = int(resumen["sin_coincidencia"])

        if conciliados == 0:
            alertas.append("No se encontraron conciliaciones concluyentes entre PSE y cruces contables.")
        if sin_coincidencia:
            alertas.append(f"{sin_coincidencia} movimientos PSE quedaron sin coincidencia en cruces contables.")
        if any(result.state == "Conciliación parcial" for result in match_results):
            alertas.append("Se detectaron conciliaciones parciales. Revisar diferencias manualmente.")

        return alertas
