import base64
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ZAPOTE_FILL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
DUPLICATE_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")


class QueryInterno:
    def __init__(
        self,
        ccs_workbook: Any,
        query_workbook: Any | None = None,
        *,
        date_tolerance_days: int = 0,
        value_tolerance: float = 0.01,
    ) -> None:
        self.ccs_workbook = ccs_workbook
        self.query_workbook = query_workbook if query_workbook is not None else ccs_workbook
        self.date_tolerance_days = max(0, int(date_tolerance_days))
        self.value_tolerance = max(0.0, float(value_tolerance))
        self.ccs_entries: List[Dict[str, Any]] = []
        self.query_entries: List[Dict[str, Any]] = []
        self.trace_rows: List[Dict[str, Any]] = []
        self.duplicate_rows: List[Dict[str, Any]] = []
        self.pending_rows: List[Dict[str, Any]] = []
        self.summary: Dict[str, int] = {
            "total_query": 0,
            "cruzados": 0,
            "validar_valor_duplicado": 0,
            "pendientes": 0,
            "ccs_utilizados": 0,
            "ccs_protegidos": 0,
            "nuevos_cruces": 0,
        }

    def _normalizar_texto(self, value: Any) -> str:
        text = str(value) if value is not None else ""
        text = text.replace("?", "o").replace("�", "o")
        normalized = unicodedata.normalize("NFKD", text)
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _normalize_account(self, value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        return text.replace(" ", "").replace("-", "").replace(".", "").lower()

    def _normalize_bank_reference(self, value: Any) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.replace("-", "").replace(" ", "").replace("_", "")
        return text

    def _resolve_ccs_sheet_for_query(self, bank_value: Any, workbook: Any) -> Optional[str]:
        bank_ref = self._normalize_bank_reference(bank_value)
        if not bank_ref:
            return None

        exact_map = {
            "bancolombiacta13321960531": "Bancolombia Cta 0531",
            "bancodebogotacta055324479": "Banco de Bogotá 0553",
            "bancolombiacta69000002490": "Bancolombia Cta 690",
            "bancolombiacta0531": "Bancolombia Cta 0531",
            "bancodebogotacta0553": "Banco de Bogotá 0553",
            "bancolombiacta690": "Bancolombia Cta 690",
        }
        for key, sheet_name in exact_map.items():
            if key in bank_ref:
                return sheet_name

        for sheet in workbook.worksheets:
            name = self._normalize_bank_reference(sheet.title)
            shared_tokens = [token for token in ("0531", "0553", "690") if token in bank_ref and token in name]
            if shared_tokens:
                return sheet.title
            if "bancolombia" in bank_ref and "bancolombia" in name:
                return sheet.title
            if "bogota" in bank_ref and "bogota" in name:
                return sheet.title
        return None

    def _matches_query_candidate(self, ccs_entry: Dict[str, Any], query_entry: Dict[str, Any]) -> bool:
        if ccs_entry["fecha"] == query_entry["fecha"] and abs(ccs_entry["valor"] - query_entry["valor"]) <= self.value_tolerance:
            return True
        delta_days = abs((ccs_entry["fecha"] - query_entry["fecha"]).days)
        if delta_days <= self.date_tolerance_days and abs(ccs_entry["valor"] - query_entry["valor"]) <= self.value_tolerance:
            return True
        return False

    def _normalize_document(self, value: Any) -> str:
        if value is None or isinstance(value, bool):
            return ""
        text = str(value).strip().upper()
        if not text:
            return ""
        cleaned = "".join(ch for ch in text if ch.isalnum())
        return cleaned.lstrip("0") if cleaned else ""

    def _parse_amount(self, value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            cleaned = str(value).strip().replace("$", "").replace(".", "").replace(",", ".")
            if not cleaned:
                return None
            return float(cleaned)
        except Exception:
            return None

    def _parse_date(self, value: Any) -> date | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in (
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%d/%m/%Y %H:%M:%S",
                "%d-%m-%Y %H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M:%S",
                "%m/%d/%Y",
                "%m/%d/%Y %H:%M:%S",
            ):
                try:
                    return datetime.strptime(text, fmt).date()
                except Exception:
                    continue
        return None

    def _ensure_annotation_columns(self, sheet: Worksheet) -> Tuple[int, int]:
        comment_col: Optional[int] = None
        observation_col: Optional[int] = None
        for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                header = self._normalizar_texto(cell.value)
                if comment_col is None and "comentario" in header:
                    comment_col = cell.column
                if observation_col is None and "observacion" in header:
                    observation_col = cell.column
        if comment_col is None:
            comment_col = sheet.max_column + 1
            sheet.cell(row=1, column=comment_col).value = "Comentario"
        if observation_col is None:
            observation_col = sheet.max_column + 1
            sheet.cell(row=1, column=observation_col).value = "Observacion"
        return comment_col, observation_col

    def _ensure_trace_columns(self, sheet: Worksheet) -> Dict[str, int]:
        columns = {
            "estado_cruce": "ESTADO_CRUCE",
            "hoja_ccs": "HOJA_CCS",
            "fuente_cruce": "FUENTE_CRUCE",
            "id_ccs_cruzado": "ID_CCS_CRUZADO",
            "fila_ccs_cruzado": "FILA_CCS_CRUZADO",
            "cuenta_ccs": "CUENTA_CCS",
            "fecha_ccs": "FECHA_CCS",
            "valor_ccs": "VALOR_CCS",
            "documento_sap_ccs": "DOCUMENTO_CCS",
            "no_documento_sap": "NO_DOCUMENTO_SAP",
            "fecha_hora_cruce": "FECHA_HORA_CRUCE",
            "resultado_cruce": "RESULTADO",
            "observacion_cruce": "OBSERVACION",
        }
        mapping: Dict[str, int] = {}
        for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                header = self._normalizar_texto(cell.value)
                for key, label in columns.items():
                    if header == self._normalizar_texto(label):
                        mapping[key] = cell.column
        for key, label in columns.items():
            if key not in mapping:
                mapping[key] = sheet.max_column + 1
                sheet.cell(row=1, column=mapping[key]).value = label
        return mapping

    def _header_map(self, sheet: Worksheet) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                header = self._normalizar_texto(cell.value)
                if not header:
                    continue
                is_account_header = (
                    "cuenta" in header
                    or "cta" in header
                    or "banco" in header
                    or "account" in header
                    or ("codigo" in header and "cliente" in header and "nombre" not in header)
                )
                if is_account_header:
                    mapping["cuenta"] = cell.column
                if "fecha" in header or "date" in header:
                    mapping["fecha"] = cell.column
                if (
                    "valor" in header
                    or "monto" in header
                    or "importe" in header
                    or "amount" in header
                    or ("recibo" in header and "importe" in header)
                ):
                    mapping["valor"] = cell.column
                if (
                    "documento" in header
                    or "sap" in header
                    or "doc" in header
                    or ("numero" in header and "documento" in header)
                ):
                    mapping["documento"] = cell.column
        if "cuenta" not in mapping and sheet.max_column >= 2:
            mapping["cuenta"] = 2
        if "fecha" not in mapping and sheet.max_column >= 4:
            mapping["fecha"] = 4
        if "valor" not in mapping and sheet.max_column >= 5:
            mapping["valor"] = 5
        if "documento" not in mapping and sheet.max_column >= 11:
            mapping["documento"] = 11
        return mapping

    def _find_state_column(self, sheet: Worksheet) -> Optional[int]:
        for row in sheet.iter_rows(min_row=1, max_row=min(15, sheet.max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                header = self._normalizar_texto(cell.value)
                if "estado" in header and "concili" in header:
                    return cell.column
        return None

    def _has_existing_state(self, sheet: Worksheet, row_idx: int) -> bool:
        state_col = self._find_state_column(sheet)
        if state_col is None:
            return False
        value = str(sheet.cell(row=row_idx, column=state_col).value or "").strip().lower()
        return any(token in value for token in ("cruzado", "conciliado", "matched", "query"))

    def _append_annotation(self, cell: Any, text: str) -> None:
        if cell is None:
            return
        current = str(cell.value or "").strip()
        if not current:
            cell.value = text
        elif text.lower() not in current.lower():
            cell.value = f"{current} | {text}"

    def _build_entry(self, sheet: Worksheet, row_idx: int, state_col: Optional[int], comment_col: Optional[int], workbook_type: str) -> Optional[Dict[str, Any]]:
        columns = self._header_map(sheet)
        if not columns:
            return None
        cuenta_col = columns.get("cuenta")
        fecha_col = columns.get("fecha")
        valor_col = columns.get("valor")
        doc_col = columns.get("documento")
        if not cuenta_col or not fecha_col or not valor_col:
            return None
        if state_col is not None and self._has_existing_state(sheet, row_idx):
            return None
        cuenta = sheet.cell(row=row_idx, column=cuenta_col).value
        fecha = sheet.cell(row=row_idx, column=fecha_col).value
        valor = sheet.cell(row=row_idx, column=valor_col).value
        doc = sheet.cell(row=row_idx, column=doc_col).value if doc_col else None
        amount = self._parse_amount(valor)
        parsed_date = self._parse_date(fecha)
        if cuenta is None or amount is None or parsed_date is None:
            return None
        normalized_doc = self._normalize_document(doc)
        return {
            "sheet": sheet,
            "row": row_idx,
            "cuenta": self._normalize_account(cuenta),
            "banco": cuenta,
            "fecha": parsed_date,
            "valor": round(amount, 2),
            "documento": str(doc).strip() if doc is not None else "",
            "documento_norm": normalized_doc,
            "state_col": state_col,
            "comment_col": comment_col,
            "kind": workbook_type,
        }

    def _extract_entries(self, workbook: Any, workbook_type: str) -> List[Dict[str, Any]]:
        entries: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            comment_col, _ = self._ensure_annotation_columns(sheet)
            state_col = self._find_state_column(sheet)
            for row_idx in range(2, sheet.max_row + 1):
                entry = self._build_entry(sheet, row_idx, state_col, comment_col, workbook_type)
                if entry is not None:
                    entries.append(entry)
        return entries

    def _set_status(self, row: Dict[str, Any], status: str) -> None:
        state_col = row.get("state_col")
        if state_col is None:
            return
        cell = row["sheet"].cell(row=row["row"], column=state_col)
        current = str(cell.value or "").strip()
        if not current:
            cell.value = status
        elif status.lower() not in current.lower():
            cell.value = f"{current} | {status}"

    def _mark_duplicate(self, query_entry: Dict[str, Any], ccs_candidates: List[Dict[str, Any]]) -> None:
        duplicate_message = (
            f"VALIDAR VALOR DUPLICADO | Cuenta {query_entry['cuenta']} | Fecha {query_entry['fecha']} | "
            f"Valor {query_entry['valor']:.2f} | Query fila {query_entry['row']}"
        )
        trace_map = self._ensure_trace_columns(query_entry["sheet"])
        self._set_query_trace(query_entry["sheet"], query_entry["row"], trace_map, {
            "estado_cruce": "VALIDAR VALOR DUPLICADO",
            "hoja_ccs": ", ".join(sorted({candidate["sheet"].title for candidate in ccs_candidates})),
            "fuente_cruce": "CCS",
            "resultado_cruce": "DUPLICADO",
            "observacion_cruce": "Existen múltiples candidatos CCS para la misma cuenta, fecha y valor",
        })
        query_cell = query_entry["sheet"].cell(row=query_entry["row"], column=query_entry["comment_col"])
        self._append_annotation(query_cell, duplicate_message)
        for candidate in ccs_candidates:
            self._append_annotation(candidate["sheet"].cell(row=candidate["row"], column=candidate["comment_col"]), duplicate_message)
            for cell in candidate["sheet"][candidate["row"]]:
                cell.fill = DUPLICATE_FILL
        self.duplicate_rows.append({
            "fila_query": query_entry["row"],
            "banco": query_entry.get("banco") or query_entry["cuenta"],
            "fecha": query_entry["fecha"].strftime("%d/%m/%Y") if isinstance(query_entry["fecha"], date) else str(query_entry["fecha"]),
            "valor": float(query_entry["valor"]),
            "sap_query": query_entry.get("documento") or "",
            "hoja_ccs": ", ".join(sorted({candidate["sheet"].title for candidate in ccs_candidates})),
            "filas_ccs_candidatas": ", ".join(str(candidate["row"]) for candidate in ccs_candidates),
            "cantidad_candidatos": len(ccs_candidates),
            "observacion": "Existen múltiples candidatos CCS para la misma cuenta, fecha y valor",
        })

    def _build_trace_message(self, ccs_row: Dict[str, Any], query_row: Dict[str, Any]) -> str:
        ccs_doc = ccs_row["documento"] or "N/A"
        query_doc = query_row["documento"] or "N/A"
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        return (
            "TRACE-QUERY | Fuente origen: CCS | Fuente conciliacion: QUERY | "
            f"ID CCS: {ccs_row['row']} | ID QUERY: {query_row['row']} | "
            f"Cuenta: {ccs_row['cuenta']} | Fecha: {ccs_row['fecha']} | "
            f"Valor: {ccs_row['valor']:.2f} | Documento SAP: {ccs_doc} | QueryDoc: {query_doc} | "
            f"Fecha/Hora: {timestamp} | Estado: CRUZADO | Resultado: MATCH"
        )

    def _set_query_trace(self, sheet: Worksheet, row_idx: int, trace_map: Dict[str, int], values: Dict[str, Any]) -> None:
        for key, value in values.items():
            if key in {"documento_sap_ccs", "no_documento_sap"}:
                for alias_key in ("documento_sap_ccs", "no_documento_sap"):
                    if alias_key not in trace_map:
                        continue
                    cell = sheet.cell(row=row_idx, column=trace_map[alias_key])
                    if value is None or value == "":
                        cell.value = "N/A" if alias_key in {"documento_sap_ccs", "no_documento_sap"} else ""
                    else:
                        cell.value = value
                continue
            if key not in trace_map:
                continue
            cell = sheet.cell(row=row_idx, column=trace_map[key])
            if value is None or value == "":
                cell.value = "N/A" if key in {"id_ccs_cruzado", "fila_ccs_cruzado", "cuenta_ccs", "fecha_ccs", "valor_ccs", "fecha_hora_cruce"} else ""
            else:
                cell.value = value

    def _apply_match(self, ccs_entry: Dict[str, Any], query_entry: Dict[str, Any]) -> None:
        trace_message = self._build_trace_message(ccs_entry, query_entry)
        self._set_status(ccs_entry, "CRUZADO")
        self._append_annotation(ccs_entry["sheet"].cell(row=ccs_entry["row"], column=ccs_entry["comment_col"]), trace_message)
        self._append_annotation(query_entry["sheet"].cell(row=query_entry["row"], column=query_entry["comment_col"]), trace_message)
        # Ensure the NO_DOCUMENTO_SAP (Documento SAP) is explicitly visible in the comment/observation
        sap_value = ccs_entry.get("documento") or "N/A"
        self._append_annotation(query_entry["sheet"].cell(row=query_entry["row"], column=query_entry["comment_col"]), f"NO_DOCUMENTO_SAP: {sap_value}")
        for cell in ccs_entry["sheet"][ccs_entry["row"]]:
            cell.fill = ZAPOTE_FILL
        for cell in query_entry["sheet"][query_entry["row"]]:
            cell.fill = ZAPOTE_FILL
        trace_map = self._ensure_trace_columns(query_entry["sheet"])
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        observacion_text = f"Cruce único por cuenta + fecha + valor + SAP | NO_DOCUMENTO_SAP: {sap_value}"
        record = {
            "fila_query": query_entry["row"],
            "banco_query": query_entry.get("banco") or query_entry["cuenta"],
            "fecha_query": query_entry["fecha"].strftime("%d/%m/%Y") if isinstance(query_entry["fecha"], date) else str(query_entry["fecha"]),
            "valor_query": float(query_entry["valor"]),
            "sap_query": query_entry.get("documento") or "",
            "hoja_ccs": ccs_entry["sheet"].title,
            "fila_ccs": ccs_entry["row"],
            "fecha_ccs": ccs_entry["fecha"].strftime("%d/%m/%Y") if isinstance(ccs_entry["fecha"], date) else str(ccs_entry["fecha"]),
            "valor_ccs": float(ccs_entry["valor"]),
            "documento_ccs": ccs_entry.get("documento") or "N/A",
            "estado": "CRUZADO",
            "resultado": "MATCH",
            "fecha_hora": timestamp,
            "observacion": observacion_text,
        }
        self.trace_rows.append(record)
        self._set_query_trace(query_entry["sheet"], query_entry["row"], trace_map, {
            "estado_cruce": "CRUZADO",
            "hoja_ccs": ccs_entry["sheet"].title,
            "fuente_cruce": "CCS",
            "id_ccs_cruzado": str(ccs_entry["row"]),
            "fila_ccs_cruzado": ccs_entry["row"],
            "cuenta_ccs": ccs_entry["cuenta"],
            "fecha_ccs": ccs_entry["fecha"].strftime("%d/%m/%Y") if isinstance(ccs_entry["fecha"], date) else str(ccs_entry["fecha"]),
            "valor_ccs": float(ccs_entry["valor"]),
            "documento_sap_ccs": ccs_entry["documento"] or "N/A",
            "no_documento_sap": ccs_entry["documento"] or "N/A",
            "fecha_hora_cruce": timestamp,
            "resultado_cruce": "MATCH",
            "observacion_cruce": observacion_text,
        })

    def _mark_pending(self, query_entry: Dict[str, Any], reason: str, result: str = "SIN MATCH") -> None:
        trace_map = self._ensure_trace_columns(query_entry["sheet"])
        self._set_query_trace(query_entry["sheet"], query_entry["row"], trace_map, {
            "estado_cruce": "PENDIENTE",
            "hoja_ccs": "N/A",
            "fuente_cruce": "CCS",
            "resultado_cruce": result,
            "observacion_cruce": reason,
        })
        self.pending_rows.append({
            "fila_query": query_entry["row"],
            "banco": query_entry.get("banco") or query_entry["cuenta"],
            "fecha": query_entry["fecha"].strftime("%d/%m/%Y") if isinstance(query_entry["fecha"], date) else str(query_entry["fecha"]),
            "valor": float(query_entry["valor"]),
            "sap": query_entry.get("documento") or "",
            "motivo": reason,
        })

    def export_final_workbook(self) -> str:
        final_wb = Workbook()
        while final_wb.sheetnames:
            final_wb.remove(final_wb[final_wb.sheetnames[0]])
        for sheet in self.query_workbook.worksheets:
            new_sheet = final_wb.create_sheet(title=sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    dst = new_sheet.cell(row=cell.row, column=cell.column)
                    dst.value = cell.value
                    if cell.number_format:
                        dst.number_format = cell.number_format
                    if cell.font:
                        dst.font = cell.font.copy()
                    if cell.alignment:
                        dst.alignment = cell.alignment.copy()
                    if cell.border:
                        dst.border = cell.border.copy()
                    if cell.fill:
                        dst.fill = cell.fill.copy()
        final_query_sheet = final_wb[self.query_workbook.worksheets[0].title]
        trace_map = self._ensure_trace_columns(final_query_sheet)
        # The status is already set in the original workbook; if the final sheet is copied from the original instance,
        # it already contains the planned values. We only ensure the required columns exist.
        summary_ws = final_wb.create_sheet(title="RESUMEN")
        summary_ws.append(["TOTAL REGISTROS QUERY", self.summary.get("total_query", 0)])
        summary_ws.append(["TOTAL CRUZADOS", self.summary.get("cruzados", 0)])
        summary_ws.append(["TOTAL VALIDAR VALOR DUPLICADO", self.summary.get("validar_valor_duplicado", 0)])
        summary_ws.append(["TOTAL PENDIENTES", self.summary.get("pendientes", 0)])
        summary_ws.append(["TOTAL SIN MATCH", self.summary.get("pendientes", 0)])
        summary_ws.append(["TOTAL CCS UTILIZADOS EN CRUCES QUERY", self.summary.get("ccs_utilizados", 0)])
        summary_ws.append(["TOTAL CCS QUE YA TENÍAN CRUCE Y FUERON PROTEGIDOS", self.summary.get("ccs_protegidos", 0)])
        summary_ws.append(["TOTAL CRUCES NUEVOS", self.summary.get("nuevos_cruces", 0)])

        trazabilidad_ws = final_wb.create_sheet(title="TRAZABILIDAD_QUERY")
        trazabilidad_ws.append([
            "FILA QUERY",
            "BANCO QUERY",
            "FECHA QUERY",
            "VALOR QUERY",
            "SAP QUERY",
            "HOJA CCS",
            "FILA CCS",
            "FECHA CCS",
            "VALOR CCS",
            "DOCUMENTO CCS",
            "NO DOCUMENTO SAP",
            "ESTADO",
            "RESULTADO",
            "FECHA/HORA",
            "OBSERVACIÓN",
        ])
        for row in self.trace_rows:
            trazabilidad_ws.append([
                row.get("fila_query"),
                row.get("banco_query"),
                row.get("fecha_query"),
                row.get("valor_query"),
                row.get("sap_query"),
                row.get("hoja_ccs"),
                row.get("fila_ccs"),
                row.get("fecha_ccs"),
                row.get("valor_ccs"),
                row.get("documento_ccs"),
                row.get("documento_ccs"),
                row.get("estado"),
                row.get("resultado"),
                row.get("fecha_hora"),
                row.get("observacion"),
            ])

        duplicados_ws = final_wb.create_sheet(title="VALIDAR_DUPLICADOS")
        duplicados_ws.append([
            "FILA QUERY",
            "BANCO",
            "FECHA",
            "VALOR",
            "SAP QUERY",
            "HOJA CCS",
            "FILAS CCS CANDIDATAS",
            "CANTIDAD CANDIDATOS",
            "OBSERVACIÓN",
        ])
        for row in self.duplicate_rows:
            duplicados_ws.append([
                row.get("fila_query"),
                row.get("banco"),
                row.get("fecha"),
                row.get("valor"),
                row.get("sap_query"),
                row.get("hoja_ccs"),
                row.get("filas_ccs_candidatas"),
                row.get("cantidad_candidatos"),
                row.get("observacion"),
            ])

        pendientes_ws = final_wb.create_sheet(title="PENDIENTES_QUERY")
        pendientes_ws.append([
            "FILA QUERY",
            "BANCO",
            "FECHA",
            "VALOR",
            "SAP",
            "MOTIVO",
        ])
        for row in self.pending_rows:
            pendientes_ws.append([
                row.get("fila_query"),
                row.get("banco"),
                row.get("fecha"),
                row.get("valor"),
                row.get("sap"),
                row.get("motivo"),
            ])

        stream = BytesIO()
        final_wb.save(stream)
        return base64.b64encode(stream.getvalue()).decode("utf-8")

    def procesar(self) -> None:
        if self.ccs_workbook is None or self.query_workbook is None:
            return
        ccs_entries = self._extract_entries(self.ccs_workbook, "ccs")
        query_entries = self._extract_entries(self.query_workbook, "query")
        if not ccs_entries or not query_entries:
            return

        self.summary["total_query"] = len(query_entries)
        used_ccs_rows: set[Tuple[str, int]] = set()
        protected_ccs = 0
        for ccs_entry in ccs_entries:
            if self._has_existing_state(ccs_entry["sheet"], ccs_entry["row"]):
                protected_ccs += 1

        for query_entry in query_entries:
            target_sheet = self._resolve_ccs_sheet_for_query(query_entry.get("banco"), self.ccs_workbook)
            if target_sheet is None:
                self._mark_pending(query_entry, "No se pudo identificar la cuenta CCS correcta para este QUERY")
                self.summary["pendientes"] += 1
                continue

            bank_candidates = [
                ccs_entry
                for ccs_entry in ccs_entries
                if ccs_entry["sheet"].title == target_sheet
                and (ccs_entry["sheet"].title, ccs_entry["row"]) not in used_ccs_rows
                and not self._has_existing_state(ccs_entry["sheet"], ccs_entry["row"])
            ]
            base_candidates = [
                ccs_entry
                for ccs_entry in bank_candidates
                if self._matches_query_candidate(ccs_entry, query_entry)
            ]
            if not base_candidates:
                self._mark_pending(query_entry, "No se encontró candidato CCS válido en la hoja correcta")
                self.summary["pendientes"] += 1
                continue

            query_doc_norm = self._normalize_document(query_entry.get("documento"))
            if query_doc_norm:
                exact_doc_candidates = [
                    ccs_entry
                    for ccs_entry in base_candidates
                    if self._normalize_document(ccs_entry.get("documento")) == query_doc_norm
                ]
                candidates = exact_doc_candidates if exact_doc_candidates else base_candidates
            else:
                candidates = base_candidates

            if len(candidates) > 1:
                unique_doc_groups = {}
                for ccs_entry in candidates:
                    key = self._normalize_document(ccs_entry.get("documento")) or "__NO_DOC__"
                    unique_doc_groups.setdefault(key, []).append(ccs_entry)
                if query_doc_norm and len(unique_doc_groups) == 1:
                    chosen = next(iter(unique_doc_groups.values()))[0]
                    used_ccs_rows.add((chosen["sheet"].title, chosen["row"]))
                    self._apply_match(chosen, query_entry)
                    self.summary["cruzados"] += 1
                    self.summary["nuevos_cruces"] += 1
                    continue
                self._mark_duplicate(query_entry, candidates)
                self.summary["validar_valor_duplicado"] += 1
                continue
            if len(candidates) == 1:
                ccs_entry = candidates[0]
                ccs_doc_norm = self._normalize_document(ccs_entry.get("documento"))
                if query_doc_norm and ccs_doc_norm and query_doc_norm != ccs_doc_norm:
                    self._append_annotation(
                        query_entry["sheet"].cell(row=query_entry["row"], column=query_entry["comment_col"]),
                        f"Documento QUERY {query_entry.get('documento')} difiere del SAP CCS {ccs_entry.get('documento')}; se conserva el cruce por cuenta + fecha + valor.",
                    )
                    self._append_annotation(
                        ccs_entry["sheet"].cell(row=ccs_entry["row"], column=ccs_entry["comment_col"]),
                        f"Documento QUERY {query_entry.get('documento')} difiere del SAP CCS {ccs_entry.get('documento')}; se conserva el cruce por cuenta + fecha + valor.",
                    )
                used_ccs_rows.add((ccs_entry["sheet"].title, ccs_entry["row"]))
                self._apply_match(ccs_entry, query_entry)
                self.summary["cruzados"] += 1
                self.summary["nuevos_cruces"] += 1
                continue

            self._mark_pending(query_entry, "No se encontró un candidato CCS inequívoco")
            self.summary["pendientes"] += 1

        self.summary["ccs_utilizados"] = len(used_ccs_rows)
        self.summary["ccs_protegidos"] = protected_ccs
        if self.summary["cruzados"] == 0 and self.summary["validar_valor_duplicado"] == 0 and self.summary["pendientes"] == 0:
            self.summary["pendientes"] = len(query_entries)
