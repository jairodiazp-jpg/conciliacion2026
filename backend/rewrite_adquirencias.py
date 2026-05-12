"""Rewrite procesador_adquirencias.py with correct implementation."""
from pathlib import Path

content = '''"""Procesador de Adquirencias: cruza valores, fechas y códigos con Bancolombia 690."""

from __future__ import annotations

import base64
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


LIGHT_BLUE_FILL = PatternFill(start_color="FFE8F4FF", end_color="FFE8F4FF", fill_type="solid")


class ProcesadorAdquirencias:
    def __init__(self, adquirencias_bytes: bytes, contable_b64: str, value_tolerance: float = 0.01) -> None:
        self.adquirencias_workbook = load_workbook(filename=BytesIO(adquirencias_bytes))
        self.contable_workbook = load_workbook(filename=BytesIO(base64.b64decode(contable_b64)))
        self.value_tolerance = value_tolerance
        self.logs: list[dict[str, Any]] = []
        self.adquirencia_counter = 1
        self.annotation_columns_adq = self._index_annotation_columns(self.adquirencias_workbook)
        self.annotation_columns_cont = self._index_annotation_columns(self.contable_workbook)

    def procesar(self) -> dict[str, str]:
        """Procesa Adquirencias y retorna dict con archivos en base64."""
        adquirencias_data = self._extraer_adquirencias_con_fila()
        self._cruzar_ambos_archivos(adquirencias_data)
        
        adq_stream = BytesIO()
        self.adquirencias_workbook.save(adq_stream)
        adq_b64 = base64.b64encode(adq_stream.getvalue()).decode("utf-8")
        
        cont_stream = BytesIO()
        self.contable_workbook.save(cont_stream)
        cont_b64 = base64.b64encode(cont_stream.getvalue()).decode("utf-8")
        
        return {
            "adquirencias_file": adq_b64,
            "contable_file": cont_b64,
        }

    def _normalizar_texto(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", str(value) if value else "")
        no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
        return no_accents.strip().lower()

    def _index_annotation_columns(self, workbook) -> dict[str, tuple[int | None, int | None]]:
        indexed: dict[str, tuple[int | None, int | None]] = {}
        for sheet in workbook.worksheets:
            comment_col: int | None = None
            observation_col: int | None = None
            max_scan_row = min(sheet.max_row, 25)
            for row in sheet.iter_rows(min_row=1, max_row=max_scan_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)
                    if comment_col is None and "coment" in header:
                        comment_col = cell.column
                    if observation_col is None and "observ" in header:
                        observation_col = cell.column
                if comment_col is not None and observation_col is not None:
                    break
            indexed[sheet.title] = (comment_col, observation_col)
        return indexed

    def _parse_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(cleaned, fmt).date()
                    except Exception:
                        continue
        return None

    def _extraer_adquirencias_con_fila(self) -> list[dict[str, Any]]:
        """Extrae datos de Adquirencias con referencias a celdas."""
        datos = []
        for sheet in self.adquirencias_workbook.worksheets:
            valor_col = None
            fecha_col = None
            auth_col = None
            max_scan = min(sheet.max_row, 25)
            header_row = 1
            
            for row_idx in range(1, max_scan + 1):
                for cell in sheet[row_idx]:
                    if not isinstance(cell.value, str):
                        continue
                    header = self._normalizar_texto(cell.value)
                    if valor_col is None and any(x in header for x in {"valor", "monto", "importe", "amount"}):
                        valor_col = cell.column
                    if fecha_col is None and any(x in header for x in {"fecha", "date", "transaccion"}):
                        fecha_col = cell.column
                    if auth_col is None and any(x in header for x in {"autorizacion", "codigo", "authorization"}):
                        auth_col = cell.column
                if valor_col and fecha_col:
                    header_row = row_idx
                    break
            
            if valor_col is None or fecha_col is None:
                continue
            
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
                try:
                    valor_cell = row[valor_col - 1] if valor_col - 1 < len(row) else None
                    fecha_cell = row[fecha_col - 1] if fecha_col - 1 < len(row) else None
                    if valor_cell is None or fecha_cell is None:
                        continue
                    
                    if isinstance(valor_cell.value, (int, float)):
                        valor = float(valor_cell.value)
                    else:
                        try:
                            valor = float(str(valor_cell.value or 0).replace(",", "."))
                        except Exception:
                            continue
                    
                    if abs(valor) < 0.0001:
                        continue
                    
                    fecha = self._parse_date(fecha_cell.value)
                    if fecha is None:
                        continue
                    
                    auth_code = ""
                    if auth_col is not None:
                        auth_cell = row[auth_col - 1] if auth_col - 1 < len(row) else None
                        if auth_cell is not None:
                            auth_code = str(auth_cell.value or "").strip()
                    
                    datos.append({
                        "sheet": sheet.title,
                        "row": row[0].row,
                        "valor": valor,
                        "fecha": fecha,
                        "autorizacion": auth_code,
                        "valor_cell": valor_cell,
                        "valor_col": valor_col,
                    })
                except Exception:
                    continue
        
        return datos

    def _cruzar_ambos_archivos(self, adquirencias_data: list[dict[str, Any]]) -> None:
        """Busca coincidencias por valor+fecha+código y colorea ambos lados."""
        bancolombia_sheet = None
        for sheet in self.contable_workbook.worksheets:
            if "1331" in sheet.title or "690" in sheet.title:
                bancolombia_sheet = sheet
                break
        
        if bancolombia_sheet is None:
            return
        
        start_row = None
        marker_text = "consignaciones sin registrar"
        for row in bancolombia_sheet.iter_rows(min_row=1, max_row=bancolombia_sheet.max_row):
            for cell in row:
                if isinstance(cell.value, str) and marker_text in self._normalizar_texto(cell.value):
                    start_row = cell.row + 1
                    break
            if start_row:
                break
        
        if start_row is None:
            start_row = 2
        
        valor_col_banco = self._find_valor_column(bancolombia_sheet)
        fecha_col_banco = self._find_fecha_column(bancolombia_sheet)
        auth_col_banco = self._find_auth_column(bancolombia_sheet)
        
        adq_by_key: dict[tuple[float, str, str], list[dict]] = {}
        for adq in adquirencias_data:
            key = (round(abs(adq["valor"]), 2), adq["fecha"].isoformat(), adq["autorizacion"])
            adq_by_key.setdefault(key, []).append(adq)
        
        matched_adq_indices = set()
        for bancolombia_row in bancolombia_sheet.iter_rows(min_row=start_row, max_row=bancolombia_sheet.max_row):
            try:
                valor_cell_banco = bancolombia_row[valor_col_banco - 1] if valor_col_banco - 1 < len(bancolombia_row) else None
                if valor_cell_banco is None:
                    continue
                
                if isinstance(valor_cell_banco.value, (int, float)):
                    banco_valor = float(valor_cell_banco.value)
                else:
                    try:
                        banco_valor = float(str(valor_cell_banco.value or 0).replace(",", "."))
                    except Exception:
                        continue
                
                if abs(banco_valor) < 0.0001:
                    continue
                
                banco_fecha = None
                if fecha_col_banco is not None:
                    fecha_cell_banco = bancolombia_row[fecha_col_banco - 1] if fecha_col_banco - 1 < len(bancolombia_row) else None
                    if fecha_cell_banco is not None:
                        banco_fecha = self._parse_date(fecha_cell_banco.value)
                
                banco_auth = ""
                if auth_col_banco is not None:
                    auth_cell_banco = bancolombia_row[auth_col_banco - 1] if auth_col_banco - 1 < len(bancolombia_row) else None
                    if auth_cell_banco is not None:
                        banco_auth = str(auth_cell_banco.value or "").strip()
                
                banco_fecha_iso = banco_fecha.isoformat() if banco_fecha else ""
                key = (round(abs(banco_valor), 2), banco_fecha_iso, banco_auth)
                
                matching_adq = adq_by_key.get(key, [])
                if not matching_adq:
                    continue
                
                adq = matching_adq[0]
                adq_idx = adquirencias_data.index(adq)
                
                if adq_idx in matched_adq_indices:
                    continue
                
                matched_adq_indices.add(adq_idx)
                
                tag = f"Adquirencia {self.adquirencia_counter}"
                self.adquirencia_counter += 1
                
                valor_cell_banco.fill = LIGHT_BLUE_FILL
                adq["valor_cell"].fill = LIGHT_BLUE_FILL
                
                self._anotar_cruce(
                    bancolombia_sheet,
                    bancolombia_row[0].row,
                    tag,
                    self.annotation_columns_cont,
                    f"Encontrado en {adq['sheet']}:fila {adq['row']}"
                )
                
                adq_sheet = self.adquirencias_workbook[adq["sheet"]]
                self._anotar_cruce(
                    adq_sheet,
                    adq["row"],
                    tag,
                    self.annotation_columns_adq,
                    f"Encontrado en {bancolombia_sheet.title}:fila {bancolombia_row[0].row}"
                )
                
                self.logs.append({
                    "tipo": "adquirencia_cruzada",
                    "valor": round(abs(banco_valor), 2),
                    "fecha": (banco_fecha.isoformat() if banco_fecha else adq["fecha"].isoformat()),
                    "confianza": 0.95,
                    "detalle": f"{tag}: coincidencia por valor {banco_valor:,.2f}, fecha {key[1]}, código {key[2]} en {bancolombia_sheet.title}:fila {bancolombia_row[0].row}",
                })
                
            except Exception:
                continue

    def _find_valor_column(self, sheet: Worksheet) -> int:
        candidates = ["valor", "monto", "importe", "amount", "consignacion"]
        for cell in sheet[1]:
            if isinstance(cell.value, str):
                h = self._normalizar_texto(cell.value)
                if any(c in h for c in candidates):
                    return cell.column
        return 3

    def _find_fecha_column(self, sheet: Worksheet) -> int | None:
        candidates = ["fecha", "date", "transaccion", "movimiento"]
        for cell in sheet[1]:
            if isinstance(cell.value, str):
                h = self._normalizar_texto(cell.value)
                if any(c in h for c in candidates):
                    return cell.column
        return None

    def _find_auth_column(self, sheet: Worksheet) -> int | None:
        candidates = ["autorizacion", "codigo", "authorization", "auth", "reference"]
        for cell in sheet[1]:
            if isinstance(cell.value, str):
                h = self._normalizar_texto(cell.value)
                if any(c in h for c in candidates):
                    return cell.column
        return None

    def _anotar_cruce(
        self,
        sheet: Worksheet,
        row_idx: int,
        tag: str,
        annotation_cols: dict[str, tuple[int | None, int | None]],
        location_text: str,
    ) -> None:
        comment_col, observation_col = annotation_cols.get(sheet.title, (None, None))
        
        if comment_col is not None:
            comment_cell = sheet.cell(row=row_idx, column=comment_col)
            comment_text = f"{tag} - Cruce con Adquirencias"
            current = str(comment_cell.value).strip() if comment_cell.value else ""
            if current and tag not in current:
                comment_cell.value = f"{current} | {comment_text}"
            elif not current:
                comment_cell.value = comment_text
        
        if observation_col is not None:
            observation_cell = sheet.cell(row=row_idx, column=observation_col)
            observation_text = f"{tag} {location_text}"
            current = str(observation_cell.value).strip() if observation_cell.value else ""
            if current and tag not in current:
                observation_cell.value = f"{current} | {observation_text}"
            elif not current:
                observation_cell.value = observation_text
'''

Path("procesador_adquirencias.py").write_text(content, encoding="utf-8")
print("OK")
