"""
Script para generar datos de prueba y ejecutar el procesador integrado.
Crea archivos Excel de ejemplo y valida que el fix funciona correctamente.
"""
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from integrador import ProcesadorIntegrado
from validacion_temporal import ValidacionTemporalConfig
import base64


def create_test_files():
    """Crea archivos de prueba realistas"""
    
    # 1. ARCHIVO PSE
    pse_wb = Workbook()
    pse_ws = pse_wb.active
    pse_ws.title = "PSE"
    pse_ws.append(["Fecha", "Descripción", "Valor", "Cuenta", "Referencia"])
    pse_ws.append(["2025-05-15", "Pago virtual PSE - Cliente ABC", 500000, "1001", "REF001"])
    pse_ws.append(["2025-05-16", "Pago virtual PSE - Cliente DEF", 750000, "1002", "REF002"])
    pse_ws.append(["2025-05-17", "Pago virtual PSE - Cliente GHI", 250000, "1001", "REF003"])
    
    pse_path = Path("tmp/PSE_PRUEBA.xlsx")
    pse_path.parent.mkdir(parents=True, exist_ok=True)
    pse_wb.save(pse_path)
    
    # 2. ARCHIVO CRUCES (aquí es donde probamos el fix - sin columna descriptiva explícita)
    cruces_wb = Workbook()
    cruces_ws = cruces_wb.active
    cruces_ws.title = "Cruces"
    cruces_ws.append(["Fecha", "Valor", "Concepto", "Referencia"])  # NO hay "Descripción"
    # Los datos PSE están en la columna "Concepto"
    cruces_ws.append(["2025-05-15", 500000, "Pago virtual PSE - Cliente ABC", "REF001"])
    cruces_ws.append(["2025-05-16", 750000, "Pago virtual PSE - Cliente DEF", "REF002"])
    cruces_ws.append(["2025-05-17", 250000, "Pago virtual PSE - Cliente GHI", "REF003"])
    
    cruces_path = Path("tmp/CRUCES_PRUEBA.xlsx")
    cruces_wb.save(cruces_path)
    
    # 3. ARCHIVO CONTABLE (tradicional)
    contable_wb = Workbook()
    contable_ws = contable_wb.active
    contable_ws.title = "Movimientos"
    contable_ws.append(["Fecha", "Descripción", "Valor", "Cuenta", "Estado", "ID Grupo"])
    # Simulamos cruce con valores que coinciden con PSE
    contable_ws.append(["2025-05-15", "Pago PSE - ABC", 500000, "1100", "", ""])
    contable_ws.append(["2025-05-16", "Pago PSE - DEF", 750000, "1100", "", ""])
    contable_ws.append(["2025-05-17", "Pago PSE - GHI", 250000, "1100", "", ""])
    
    contable_path = Path("tmp/CONTABLE_PRUEBA.xlsx")
    contable_wb.save(contable_path)
    
    return pse_path, cruces_path, contable_path


def run_test():
    """Ejecuta el procesador integrado con los datos de prueba"""
    
    print("=" * 70)
    print("TEST LOCAL: Procesador Integrado con Fix PSE")
    print("=" * 70)
    
    # Crear archivos de prueba
    print("\n1️⃣  Generando archivos de prueba...")
    pse_path, cruces_path, contable_path = create_test_files()
    print(f"   ✓ PSE: {pse_path}")
    print(f"   ✓ CRUCES: {cruces_path}")
    print(f"   ✓ CONTABLE: {contable_path}")
    
    # Leer archivos
    print("\n2️⃣  Leyendo archivos...")
    pse_bytes = pse_path.read_bytes()
    cruces_bytes = cruces_path.read_bytes()
    contable_bytes = contable_path.read_bytes()
    print("   ✓ Archivos leídos correctamente")
    
    # Procesar
    print("\n3️⃣  Procesando archivos...")
    try:
        temporal_config = ValidacionTemporalConfig(
            tolerancia_dias=0,
            permitir_cruce_mes_anterior=False,
            permitir_cruce_ano_anterior=False,
        )
        
        processor = ProcesadorIntegrado(
            contable_bytes=contable_bytes,
            pse_bytes=pse_bytes,
            cruces_bytes=cruces_bytes,
            date_tolerance_days=1,
            value_tolerance=0.01,
            temporal_config=temporal_config,
        )
        
        result = processor.procesar()
        print("   ✓ Procesamiento completado")
        
    except Exception as e:
        print(f"   ❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Guardar resultados (archivos listados en result['files'])
    print("\n4️⃣  Guardando archivos de resultado...")
    out_dir = Path("tmp/resultados")
    out_dir.mkdir(parents=True, exist_ok=True)

    files_saved = []
    for entry in result.get("files", []):
        name = entry.get("name") or "output.xlsx"
        data = entry.get("file")
        if isinstance(data, str):
            data = base64.b64decode(data)
        filepath = out_dir / name
        filepath.write_bytes(data)
        files_saved.append(filepath)
        print(f"   ✓ {name}")
    
    # Mostrar resumen de resultados
    print("\n5️⃣  Resumen de procesamiento:")
    print(f"\n   Modo: {result.get('mode', 'N/A')}")
    
    if "resumen" in result:
        resumen = result["resumen"]
        print(f"\n   CONTABLE:")
        print(f"   - Total movimientos: {resumen.get('contable', {}).get('total_movimientos', 0)}")
        print(f"   - Conciliados: {resumen.get('contable', {}).get('conciliados', 0)}")
        print(f"   - Parciales: {resumen.get('contable', {}).get('parciales', 0)}")
        print(f"   - Sin coincidencia: {resumen.get('contable', {}).get('sin_coincidencia', 0)}")
        
        print(f"\n   PSE:")
        print(f"   - Total movimientos: {resumen.get('pse', {}).get('total_movimientos', 0)}")
        print(f"   - Conciliados: {resumen.get('pse', {}).get('conciliados', 0)}")
        print(f"   - Sin coincidencia: {resumen.get('pse', {}).get('sin_coincidencia', 0)}")
    
    # Validación crítica: dataset_cruces debe estar lleno
    print(f"\n6️⃣  Validación del Fix:")
    dataset_cruces = result.get("pse", {}).get("dataset_cruces", [])
    print(f"   - dataset_cruces registros: {len(dataset_cruces)}")
    
    if len(dataset_cruces) > 0:
        print(f"   ✅ ÉXITO: Cruces PSE detectados y marcados")
        for i, row in enumerate(dataset_cruces[:3], 1):
            print(f"      {i}. {row.get('sheet')}:{row.get('row')} Estado={row.get('estado_conciliacion')}")
    else:
        print(f"   ❌ FALLO: No hay cruces PSE marcados")
        return False
    
    # Logs
    logs = result.get("logs", [])
    if logs:
        print(f"\n7️⃣  Logs ({len(logs)} eventos):")
        for log in logs[:5]:
            print(f"   - {log.get('tipo')}: {log.get('detalle')[:60]}...")
    
    # Alertas
    alertas = result.get("alertas", [])
    if alertas:
        print(f"\n   ⚠️  ALERTAS ({len(alertas)}):")
        for alerta in alertas[:3]:
            print(f"   - {alerta}")
    
    print("\n" + "=" * 70)
    print("✅ TEST COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"\nArchivos guardados en: {out_dir.absolute()}")
    
    return True


if __name__ == "__main__":
    import sys
    try:
        success = run_test()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
