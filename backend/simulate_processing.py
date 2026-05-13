"""Simulación paso a paso del procesamiento."""

import base64
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from procesador_adquirencias import ProcesadorAdquirencias

base = Path('..') / 'tmp_e2e'
adq_path = base / 'ADQUIRENCIAS MAYO (1).xlsx'
memo_path = base / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx'

print("=== SIMULACIÓN PASO A PASO ===\n")

# 1. Cargar archivos
adq_bytes = adq_path.read_bytes()
memo_bytes = memo_path.read_bytes()
memo_b64 = base64.b64encode(memo_bytes).decode('utf-8')

# 2. Crear procesador
processor = ProcesadorAdquirencias(adq_bytes, memo_b64)

# 3. Extraer datos y cruzar
adq_data = processor._extraer_adquirencias_con_fila()
print(f"Adquirencias extraídas: {len(adq_data)}")

# 4. Ejecutar cruce
print("\nEjecutando cruce...")
processor._cruzar_ambos_archivos(adq_data)
print(f"Contadores: {processor.adquirencia_counter} cruces encontrados")

# 5. Verificar hojas después del cruce
sheet_690 = processor.contable_workbook['Bancolombia Cta 690 ']
print(f"\nHoja 690 después del cruce:")
print(f"  Max column: {sheet_690.max_column}")
print(f"  Row 19 column 17: '{sheet_690.cell(19, 17).value}'")
print(f"  Row 19 column 18: '{sheet_690.cell(19, 18).value}'")

# 6. Guardar a bytes y recargar para simular lo que envía el endpoint
print("\nGuardando y recargando...")
output = BytesIO()
processor.contable_workbook.save(output)
output_bytes = output.getvalue()

output_wb = load_workbook(BytesIO(output_bytes))
output_sheet = output_wb['Bancolombia Cta 690 ']
print(f"Después de guardar/recargar:")
print(f"  Row 19 column 17: '{output_sheet.cell(19, 17).value}'")
print(f"  Row 19 column 18: '{output_sheet.cell(19, 18).value}'")

# 7. Verificar si hay comentarios en las primeras 20 filas
comment_count = 0
for row in range(2, 21):
    cell = output_sheet.cell(row, 17)
    if cell.value and 'Adquirencia' in str(cell.value):
        comment_count += 1
        print(f"  Row {row}: '{cell.value}'")

print(f"\nTotal comentarios encontrados: {comment_count}")
