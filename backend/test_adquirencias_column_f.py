"""Smoke test para verificar extracción de código de aprobación desde columna F del 690."""

import base64
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from procesador_adquirencias import ProcesadorAdquirencias

def crear_contable_690():
    """Crea un archivo Contable simulado con sección 690."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bancolombia 690"
    
    # Headers
    ws['A1'] = "Descripcion"
    ws['B1'] = "Fecha"
    ws['C1'] = "Concepto"
    ws['D1'] = "Valor"
    ws['E1'] = "Saldo"
    ws['F1'] = "Numero Aprobacion"
    
    # Marker row
    ws['A2'] = "consignaciones sin registrar"
    
    # Data rows
    ws['A3'] = "Transaccion"
    ws['B3'] = "01/05/2025"
    ws['C3'] = "Consignacion Adquirencias"
    ws['D3'] = 5000.00
    ws['E3'] = 5000.00  # Saldo (mismo que D3 para que coincida el heurístico)
    ws['F3'] = "166580"  # Codigo de aprobacion en columna F
    
    ws['A4'] = "Transaccion"
    ws['B4'] = "02/05/2025"
    ws['C4'] = "Otra consignacion"
    ws['D4'] = 3000.00
    ws['E4'] = 3000.00  # Saldo (mismo que D4)
    ws['F4'] = "999999"  # Otro codigo que no debe coincidir
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def crear_adquirencias():
    """Crea un archivo Adquirencias simulado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Headers (mínimo: fecha, valor, código autorización)
    ws['A1'] = "Col1"
    ws['B1'] = "Fecha Transaccion"
    ws['C1'] = "Col3"
    # ... columnas dummy hasta la 16
    for col in range(4, 16):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['P1'] = "Valor Total"  # Columna 16
    # ... más columnas
    for col in range(17, 23):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['W1'] = "Codigo Autorizacion"  # Columna 23
    
    # Data rows
    ws['B2'] = "01/05/2025"
    ws['P2'] = 5000.00
    ws['W2'] = "166580"  # Mismo código que el 690
    
    ws['B3'] = "02/05/2025"
    ws['P3'] = 2000.00
    ws['W3'] = "111111"  # Código que no coincide
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def main():
    print("=== Smoke Test: Extracción de Código de Aprobación desde Columna F ===\n")
    
    # Crear archivos
    contable_bytes = crear_contable_690()
    adquirencias_bytes = crear_adquirencias()
    
    # Codificar contable a base64
    contable_b64 = base64.b64encode(contable_bytes).decode('utf-8')
    
    # Procesar
    print("1. Procesando archivos...")
    processor = ProcesadorAdquirencias(
        adquirencias_bytes=adquirencias_bytes,
        contable_b64=contable_b64,
        value_tolerance=0.01,
        date_tolerance_days=0
    )
    
    result = processor.procesar()
    
    # Decodificar resultados
    from openpyxl import load_workbook
    
    adq_bytes = base64.b64decode(result['adquirencias_file'])
    cont_bytes = base64.b64decode(result['contable_file'])
    
    adq_wb = load_workbook(BytesIO(adq_bytes))
    cont_wb = load_workbook(BytesIO(cont_bytes))
    
    # Verificar colores en Adquirencias
    adq_ws = adq_wb['Hoja1']
    adq_p2_color = adq_ws['P2'].fill.start_color.rgb if adq_ws['P2'].fill.start_color else None
    print(f"\n2. Color en Adquirencias P2 (Valor 5000, Auth 166580): {adq_p2_color}")
    print(f"   Esperado: FFE8F4FF (azul claro)")
    
    # Verificar colores en Contable 690
    cont_ws = cont_wb['Bancolombia 690']
    cont_d3_color = cont_ws['D3'].fill.start_color.rgb if cont_ws['D3'].fill.start_color else None
    print(f"\n3. Color en Contable D3 (Valor 5000, Auth en F3=166580): {cont_d3_color}")
    print(f"   Esperado: FFE8F4FF (azul claro)")
    
    # Verificar anotaciones
    cont_comment_col = None
    for col in range(1, cont_ws.max_column + 1):
        header = cont_ws.cell(row=1, column=col).value
        if header and 'coment' in str(header).lower():
            cont_comment_col = col
            break
    
    if cont_comment_col:
        comment_val = cont_ws.cell(row=3, column=cont_comment_col).value
        print(f"\n4. Anotación en Contable fila 3: {comment_val}")
        print(f"   Esperado: Contiene 'Adquirencia 1' y 'Cruce con Adquirencias'")
    else:
        print("\n4. Anotación: No se encontró columna de comentario")
    
    # Verificar logs
    print(f"\n5. Logs generados:")
    for log in processor.logs:
        if log['tipo'] == 'adquirencia_cruzada':
            print(f"   - {log['tipo']}: {log['detalle']}")
    
    # Resultado final
    success = (
        adq_p2_color == 'FFE8F4FF' and 
        cont_d3_color == 'FFE8F4FF' and
        len(processor.logs) > 0
    )
    
    print(f"\n{'✓' if success else '✗'} TEST {'PASÓ' if success else 'FALLÓ'}")
    return 0 if success else 1

if __name__ == '__main__':
    exit(main())
