from openpyxl import Workbook
from agrupacion_pse import conciliacion_por_agrupacion


def find_max_grp_number_from_sheet(ws):
    # simple scan for existing GRP-XXXX
    maxn = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.startswith('GRP-'):
                try:
                    num = int(cell.split('-')[1])
                    if num > maxn:
                        maxn = num
                except Exception:
                    pass
    return maxn


def marca_pse_predicate(ws, row_idx):
    # deja que la columna 2 (descripcion) contenga 'PSE'
    val = ws.cell(row=row_idx, column=2).value
    return isinstance(val, str) and 'PSE' in val.upper()


def run_test():
    wb = Workbook()
    ws = wb.active
    # columnas: 1:ID, 2:Descripcion, 3:Valor, 4:Estado, 5:ID_Grupo, 6:Comentario
    ws.append(['ID','Descripcion','Valor','Estado','ID_Grupo','Comentario'])
    # total marcado PSE
    ws.append([1, 'TOTAL PSE', 30125000, None, None, None])
    # desgloses
    ws.append([2, 'detalle A', 10000000, None, None, None])
    ws.append([3, 'detalle B', 20000000, None, None, None])
    ws.append([4, 'detalle C', 125000, None, None, None])

    # pse_dataset con los valores
    pse_dataset = [
        {'valor': 10000000},
        {'valor': 20000000},
        {'valor': 125000},
    ]

    col_map = {'valor':3,'estado':4,'id_grupo':5,'comentario':6,'descripcion':2}

    conciliacion_por_agrupacion(ws, pse_dataset, col_map, lambda: find_max_grp_number_from_sheet(ws), marca_pse_predicate)

    # imprimir resultados
    for i in range(2, ws.max_row+1):
        row = [ws.cell(row=i, column=c).value for c in range(1,7)]
        print(row)


if __name__ == '__main__':
    run_test()
