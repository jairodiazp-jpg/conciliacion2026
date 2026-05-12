"""Debug detallado de extracción del 690."""

import base64
from datetime import date
from io import BytesIO
from openpyxl import Workbook, load_workbook

from procesador_adquirencias import ProcesadorAdquirencias

def crear_contable_690():
    """Crea un archivo Contable simulado con sección 690."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bancolombia 690"
    
    # Headers
    ws['A1'] = "Descripcion"
    ws['B1'] = "Fecha"
    ws['C1'] = "Concepto"
    ws['D1'] = "Valor"
    ws['E1'] = "Saldo"
    ws['F1'] = "Numero Aprobacion"
    
    # Marker row
    ws['A2'] = "consignaciones sin registrar"
    
    # Data rows
    ws['A3'] = "Transaccion"
    ws['B3'] = "01/05/2025"
    ws['C3'] = "Consignacion Adquirencias"
    ws['D3'] = 5000.00
    ws['E3'] = 5000.00  
    ws['F3'] = "166580"  
    
    ws['A4'] = "Transaccion"
    ws['B4'] = "02/05/2025"
    ws['C4'] = "Otra consignacion"
    ws['D4'] = 3000.00
    ws['E4'] = 3000.00  
    ws['F4'] = "999999"  
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def crear_adquirencias():
    """Crea un archivo Adquirencias simulado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    ws['A1'] = "Col1"
    ws['B1'] = "Fecha Transaccion"
    ws['C1'] = "Col3"
    for col in range(4, 16):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['P1'] = "Valor Total"  
    for col in range(17, 23):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['W1'] = "Codigo Autorizacion"  
    
    ws['B2'] = "01/05/2025"
    ws['P2'] = 5000.00
    ws['W2'] = "166580"  
    
    ws['B3'] = "02/05/2025"
    ws['P3'] = 2000.00
    ws['W3'] = "111111"  
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def main():
    print("=== Debug: Investigar extracción del 690 ===\n")
    
    contable_bytes = crear_contable_690()
    adquirencias_bytes = crear_adquirencias()
    contable_b64 = base64.b64encode(contable_bytes).decode('utf-8')
    
    processor = ProcesadorAdquirencias(
        adquirencias_bytes=adquirencias_bytes,
        contable_b64=contable_b64,
        value_tolerance=0.01,
        date_tolerance_days=0
    )
    
    # Test _find_header_columns
    contable_wb = processor.contable_workbook
    sheet_690 = None
    for candidate in contable_wb.worksheets:
        if "690" in candidate.title.lower():
            sheet_690 = candidate
            break
    
    print(f"1. Sheet encontrada: {sheet_690.title if sheet_690 else 'NO ENCONTRADA'}")
    
    # Test detección del marcador
    print("\n2. Buscando marcador 'consignaciones sin registrar'...")
    for row_idx, row in enumerate(sheet_690.iter_rows(min_row=1, max_row=sheet_690.max_row), start=1):
        texts = [str(cell.value) for cell in row if cell.value]
        full_text = " ".join(texts).lower()
        if "consignaciones sin registrar" in full_text:
            print(f"   Encontrado en fila {row_idx}")
            start_row = row_idx + 1
            break
    else:
        print("   NO ENCONTRADO")
        start_row = 2
    
    print(f"   Start row para datos: {start_row}")
    
    # Test _find_header_columns
    print("\n3. Buscando encabezados con _find_header_columns(start_row={})...".format(start_row - 1 if start_row > 1 else 1))
    valor_col, fecha_col, auth_col, header_row = processor._find_header_columns(
        sheet_690, 
        start_row=start_row - 1 if start_row > 1 else 1,
        need_auth=False
    )
    print(f"   valor_col={valor_col}, fecha_col={fecha_col}, auth_col={auth_col}, header_row={header_row}")
    
    # Test _parse_value_candidates en fila 3
    print("\n4. Probando _parse_value_candidates en fila 3...")
    row_3 = list(sheet_690.iter_rows(min_row=3, max_row=3, min_col=1, max_col=sheet_690.max_column))[0]
    candidates = processor._parse_value_candidates(list(row_3), exclude_cols={6})
    print(f"   Candidatos: {[(col, val) for col, val, _ in candidates]}")
    if candidates:
        min_val = min(candidates, key=lambda x: abs(x[1]))
        max_val = max(candidates, key=lambda x: abs(x[1]))
        print(f"   Min: col={min_val[0]}, val={min_val[1]}")
        print(f"   Max: col={max_val[0]}, val={max_val[1]}")
    
    # Test extracción completa
    print("\n5. Extrayendo con _extraer_movimientos_690()...")
    sheet_690_result, entries = processor._extraer_movimientos_690()
    print(f"   Sheet: {sheet_690_result.title if sheet_690_result else 'NO'}")
    print(f"   Entries: {len(entries)}")
    for entry in entries:
        print(f"     - Row {entry['row']}: valor={entry['valor']}, fecha={entry['fecha']}, auth='{entry['autorizacion']}'")

if __name__ == '__main__':
    main()
