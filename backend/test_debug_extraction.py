"""Test de depuración para verificar extracción de datos."""

import base64
from datetime import date
from io import BytesIO
from openpyxl import Workbook, load_workbook

from procesador_adquirencias import ProcesadorAdquirencias

def crear_contable_690():
    """Crea un archivo Contable simulado con sección 690."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bancolombia 690"
    
    # Headers
    ws['A1'] = "Descripcion"
    ws['B1'] = "Fecha"
    ws['C1'] = "Concepto"
    ws['D1'] = "Valor"
    ws['E1'] = "Saldo"
    ws['F1'] = "Numero Aprobacion"
    
    # Marker row
    ws['A2'] = "consignaciones sin registrar"
    
    # Data rows
    ws['A3'] = "Transaccion"
    ws['B3'] = "01/05/2025"
    ws['C3'] = "Consignacion Adquirencias"
    ws['D3'] = 5000.00  # Valor de consignación
    ws['E3'] = 5000.00  # Saldo (mismo que D3)
    ws['F3'] = "166580"  # Codigo de aprobacion en columna F
    
    ws['A4'] = "Transaccion"
    ws['B4'] = "02/05/2025"
    ws['C4'] = "Otra consignacion"
    ws['D4'] = 3000.00
    ws['E4'] = 3000.00  # Saldo (mismo que D4)
    ws['F4'] = "999999"  # Otro codigo que no debe coincidir
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def crear_adquirencias():
    """Crea un archivo Adquirencias simulado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Headers (mínimo: fecha, valor, código autorización)
    ws['A1'] = "Col1"
    ws['B1'] = "Fecha Transaccion"
    ws['C1'] = "Col3"
    # ... columnas dummy hasta la 16
    for col in range(4, 16):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['P1'] = "Valor Total"  # Columna 16
    # ... más columnas
    for col in range(17, 23):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['W1'] = "Codigo Autorizacion"  # Columna 23
    
    # Data rows
    ws['B2'] = "01/05/2025"
    ws['P2'] = 5000.00
    ws['W2'] = "166580"  # Mismo código que el 690
    
    ws['B3'] = "02/05/2025"
    ws['P3'] = 2000.00
    ws['W3'] = "111111"  # Código que no coincide
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def main():
    print("=== Test de Depuración: Extracción de Datos ===\n")
    
    # Crear archivos
    contable_bytes = crear_contable_690()
    adquirencias_bytes = crear_adquirencias()
    
    # Codificar contable a base64
    contable_b64 = base64.b64encode(contable_bytes).decode('utf-8')
    
    # Procesar
    print("1. Creando procesador...")
    processor = ProcesadorAdquirencias(
        adquirencias_bytes=adquirencias_bytes,
        contable_b64=contable_b64,
        value_tolerance=0.01,
        date_tolerance_days=0
    )
    
    # Extraer datos de Adquirencias
    print("\n2. Extrayendo datos de Adquirencias...")
    adq_data = processor._extraer_adquirencias_con_fila()
    for adq in adq_data:
        print(f"   - Sheet: {adq['sheet']}, Row: {adq['row']}")
        print(f"     Valor: {adq['valor']}, Fecha: {adq['fecha']}, Auth: '{adq['autorizacion']}'")
    
    # Extraer datos de 690
    print("\n3. Extrayendo datos de 690...")
    sheet_690, banco_data = processor._extraer_movimientos_690()
    if sheet_690:
        print(f"   Sheet: {sheet_690.title}")
        for bank in banco_data:
            print(f"   - Row: {bank['row']}")
            print(f"     Valor: {bank['valor']}, Fecha: {bank['fecha']}, Auth: '{bank['autorizacion']}'")
    else:
        print("   No se encontró sheet 690")
    
    # Procesar
    print("\n4. Ejecutando cruce...")
    result = processor.procesar()
    
    print(f"\n5. Logs generados: {len(processor.logs)}")
    for log in processor.logs:
        print(f"   - {log}")

if __name__ == '__main__':
    main()
