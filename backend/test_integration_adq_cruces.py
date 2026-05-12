"""Test de integración: Adquirencias cruza con Memorando/Cruces Contables (890)."""

import base64
from io import BytesIO
from openpyxl import Workbook, load_workbook

from integrador import ProcesadorIntegrado

def crear_cruces_contables_690():
    """Simula un memorando/cruces contables con varias hojas, incluida 690."""
    wb = Workbook()
    
    # Hoja 1: Resumen o introducción
    ws_intro = wb.active
    ws_intro.title = "Resumen"
    ws_intro['A1'] = "Memorando de Conciliación"
    ws_intro['A2'] = "Cruces Contables - Bancolombia"
    
    # Hoja 2: Cuenta 690 (la que importa)
    ws_690 = wb.create_sheet("Bancolombia 690")
    ws_690['A1'] = "Descripcion"
    ws_690['B1'] = "Fecha"
    ws_690['C1'] = "Concepto"
    ws_690['D1'] = "Valor"
    ws_690['E1'] = "Saldo"
    ws_690['F1'] = "Numero Aprobacion"
    
    # Marker
    ws_690['A2'] = "consignaciones sin registrar"
    
    # Datos
    ws_690['A3'] = "Transaccion"
    ws_690['B3'] = "01/05/2025"
    ws_690['C3'] = "Consignacion Adquirencias"
    ws_690['D3'] = 12500.00
    ws_690['E3'] = 12500.00
    ws_690['F3'] = "166580"
    
    ws_690['A4'] = "Transaccion"
    ws_690['B4'] = "02/05/2025"
    ws_690['C4'] = "Otra consignacion"
    ws_690['D4'] = 8900.00
    ws_690['E4'] = 8900.00
    ws_690['F4'] = "155555"
    
    # Hoja 3: Otra cuenta (para simular estructura real)
    ws_otra = wb.create_sheet("Otro Banco")
    ws_otra['A1'] = "Cuenta"
    ws_otra['B1'] = "Fecha"
    ws_otra['C1'] = "Valor"
    ws_otra['A2'] = "123456789"
    ws_otra['B2'] = "05/05/2025"
    ws_otra['C2'] = 5000.00
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def crear_adquirencias():
    """Simula documento de Adquirencias."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja2"
    
    # Headers
    ws['A1'] = "Col1"
    ws['B1'] = "Fecha de Transaccion"
    ws['C1'] = "Col3"
    for col in range(4, 16):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['P1'] = "Valor Total"
    for col in range(17, 23):
        ws.cell(row=1, column=col).value = f"Col{col}"
    ws['W1'] = "Codigo Autorizacion"
    
    # Data
    ws['B2'] = "01/05/2025"
    ws['P2'] = 12500.00
    ws['W2'] = "166580"  # Coincide con 690
    
    ws['B3'] = "02/05/2025"
    ws['P3'] = 8900.00
    ws['W3'] = "155555"  # Coincide con 690
    
    ws['B4'] = "03/05/2025"
    ws['P4'] = 3000.00
    ws['W4'] = "999999"  # NO coincide
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def main():
    print("=== Test de Integración: Adquirencias + Memorando Contable 690 ===\n")
    
    # Crear archivos
    cruces_bytes = crear_cruces_contables_690()
    adquirencias_bytes = crear_adquirencias()
    
    print("1. Archivos creados:")
    print(f"   - Memorando Cruces: {len(cruces_bytes)} bytes")
    print(f"   - Adquirencias: {len(adquirencias_bytes)} bytes")
    
    # Procesar
    print("\n2. Procesando con ProcesadorIntegrado...")
    processor = ProcesadorIntegrado(
        contable_bytes=None,
        pse_bytes=None,
        cruces_bytes=cruces_bytes,
        adquirencias_bytes=adquirencias_bytes,
        date_tolerance_days=0,
        value_tolerance=0.01,
    )
    
    try:
        result = processor.procesar()
        print(f"   Modo: {result.get('mode', 'unknown')}")
        print(f"   Logs: {len(result.get('logs', []))} entradas")
        print(f"   Archivos: {len(result.get('files', []))} archivos")
        
        # Verificar logs de Adquirencias
        adq_logs = [log for log in result.get('logs', []) if 'adquirencia' in str(log).lower()]
        print(f"\n3. Logs de Adquirencias encontrados: {len(adq_logs)}")
        for log in adq_logs:
            print(f"   - {log}")
        
        # Verificar archivos retornados
        print(f"\n4. Archivos generados:")
        for file_info in result.get('files', []):
            print(f"   - {file_info['name']}")
        
        # Decodificar y verificar cruces procesados
        print(f"\n5. Verificando colores en memorando de cruces...")
        for file_info in result.get('files', []):
            if 'CRUCES' in file_info['name'].upper():
                cruces_result_bytes = base64.b64decode(file_info['file'])
                cruces_wb = load_workbook(BytesIO(cruces_result_bytes))
                ws_690 = cruces_wb['Bancolombia 690']
                
                # Verificar colores en filas 3 y 4
                color_d3 = ws_690['D3'].fill.start_color.rgb if ws_690['D3'].fill.start_color else None
                color_d4 = ws_690['D4'].fill.start_color.rgb if ws_690['D4'].fill.start_color else None
                
                print(f"   Color en D3 (12500, auth 166580): {color_d3}")
                print(f"   Color en D4 (8900, auth 155555): {color_d4}")
                
                # Verificar comentarios
                comment_col = None
                for col in range(1, ws_690.max_column + 1):
                    header = ws_690.cell(row=1, column=col).value
                    if header and 'coment' in str(header).lower():
                        comment_col = col
                        break
                
                if comment_col:
                    comment_d3 = ws_690.cell(row=3, column=comment_col).value
                    comment_d4 = ws_690.cell(row=4, column=comment_col).value
                    print(f"\n   Comentario D3: {comment_d3}")
                    print(f"   Comentario D4: {comment_d4}")
        
        print(f"\n✓ TEST COMPLETADO")
        
    except Exception as e:
        print(f"✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
