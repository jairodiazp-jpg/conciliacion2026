from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pse_conciliador import PseConciliador

def test_pse_combination_and_coloring():
    # 1. Crear archivo PSE con múltiples movimientos que suman un total
    pse_wb = Workbook()
    pse_ws = pse_wb.active
    pse_ws.title = "Movimientos PSE"
    pse_ws.append(["Fecha", "Valor", "Descripción"])
    # Estos 4 suman 10,000,000
    pse_ws.append(["2025-05-15", 2500000, "Pago PSE 1"])
    pse_ws.append(["2025-05-15", 1500000, "Pago PSE 2"])
    pse_ws.append(["2025-05-15", 3000000, "Pago PSE 3"])
    pse_ws.append(["2025-05-15", 3000000, "Pago PSE 4"])
    
    # 2. Crear archivo de Cruces (CCS Memorando) con un único registro de 10,000,000
    cruces_wb = Workbook()
    cruces_ws = cruces_wb.active
    cruces_ws.title = "Bancolombia 0531"
    cruces_ws.append(["Fecha", "Valor", "Descripción"])
    cruces_ws.append(["2025-05-15", 10000000, "Pago virtual PSE"])
    
    pse_bytes = BytesIO()
    pse_wb.save(pse_bytes)
    
    cruces_bytes = BytesIO()
    cruces_wb.save(cruces_bytes)
    
    # 3. Procesar con el conciliador
    conciliador = PseConciliador(pse_bytes.getvalue(), cruces_bytes.getvalue())
    resultado = conciliador.procesar()
    
    # 4. Verificar resultados
    print("=" * 60)
    print("TEST: Combinación PSE (1 CCS -> N PSE) y Marcación Color")
    print("=" * 60)
    
    # El resultado['secondary_file'] es el de Cruces (CCS)
    import base64
    from openpyxl import load_workbook
    
    res_cruces_wb = load_workbook(BytesIO(base64.b64decode(resultado['secondary_file'])))
    res_pse_wb = load_workbook(BytesIO(base64.b64decode(resultado['file'])))
    
    pink_hex = "FFFFC0CB"
    
    # Verificar color en CCS
    ccs_sheet = res_cruces_wb["Bancolombia 0531"]
    ccs_val_cell = ccs_sheet.cell(row=2, column=2) # Columna Valor
    ccs_color = ccs_val_cell.fill.start_color.index
    
    print(f"Color celda CCS (Fila 2, Col 2): {ccs_color}")
    
    # Verificar color en PSE
    pse_sheet = res_pse_wb["Movimientos PSE"]
    all_pse_pink = True
    for r in range(2, 6):
        cell = pse_sheet.cell(row=r, column=2)
        color = cell.fill.start_color.index
        print(f"Color celda PSE (Fila {r}, Col 2): {color}")
        if color != pink_hex:
            all_pse_pink = False
            
    success = True
    if ccs_color == pink_hex:
        print("\n✅ ÉXITO: El registro en CCS Memorando está marcado en ROSADO")
    else:
        print("\n❌ FALLO: El registro en CCS Memorando NO está marcado en ROSADO")
        success = False
        
    if all_pse_pink:
        print("✅ ÉXITO: Todos los movimientos en PSE están marcados en ROSADO")
    else:
        print("❌ FALLO: Algunos movimientos en PSE NO están marcados en ROSADO")
        success = False
        
    if success:
        print("\n" + "=" * 60)
        print("RESULTADO FINAL: ✅ LÓGICA Y COLOR VALIDADO")
        print("=" * 60)
    return success

if __name__ == "__main__":
    if test_pse_combination_and_coloring():
        exit(0)
    else:
        exit(1)