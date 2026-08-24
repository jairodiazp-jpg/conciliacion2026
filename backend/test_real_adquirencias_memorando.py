"""Test con archivos REALES: Adquirencias + CCS Memorando con cuenta 690."""

import base64
import sys
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

# Agregar backend al path
sys.path.insert(0, str(Path(__file__).parent))

from integrador import ProcesadorIntegrado

def main():
    print("=== Test REAL: Adquirencias MAYO + CCS Memorando Bancolombia 690 ===\n")
    
    # Rutas de archivos reales
    adquirencias_path = Path(__file__).parent.parent / "tmp_e2e" / "ADQUIRENCIAS MAYO (1).xlsx"
    memorando_path = Path(__file__).parent.parent / "tmp_e2e" / "CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx"
    
    print("1. Verificando archivos...")
    if not adquirencias_path.exists():
        print(f"   ✗ NO ENCONTRADO: {adquirencias_path}")
        return
    if not memorando_path.exists():
        print(f"   ✗ NO ENCONTRADO: {memorando_path}")
        return
    
    print(f"   ✓ Adquirencias: {adquirencias_path.name} ({adquirencias_path.stat().st_size} bytes)")
    print(f"   ✓ Memorando: {memorando_path.name} ({memorando_path.stat().st_size} bytes)")
    
    # Leer archivos
    print("\n2. Leyendo archivos...")
    with open(adquirencias_path, 'rb') as f:
        adquirencias_bytes = f.read()
    with open(memorando_path, 'rb') as f:
        memorando_bytes = f.read()
    
    # Inspeccionar estructura del memorando
    print("\n3. Inspeccionando estructura del memorando...")
    memorando_wb = load_workbook(memorando_path)
    for sheet_name in memorando_wb.sheetnames:
        ws = memorando_wb[sheet_name]
        print(f"   Hoja: {sheet_name}")
        if '690' in sheet_name.lower() or 'bancolombia' in sheet_name.lower():
            print(f"      → Contiene cuenta 690 ✓")
            # Mostrar headers
            headers = [cell.value for cell in ws[1]]
            print(f"      → Columnas detectadas: {headers[:8]}...")
    
    # Procesar con integrador
    print("\n4. Procesando cruce Adquirencias + Memorando...")
    try:
        processor = ProcesadorIntegrado(
            contable_bytes=None,
            pse_bytes=None,
            cruces_bytes=memorando_bytes,
            adquirencias_bytes=adquirencias_bytes,
            date_tolerance_days=2,
            value_tolerance=0.01,
        )
        
        result = processor.procesar()
        
        print(f"   ✓ Modo: {result.get('mode')}")
        print(f"   ✓ Logs: {len(result.get('logs', []))} entradas")
        print(f"   ✓ Alertas: {len(result.get('alertas', []))}")
        
        # Filtrar logs de adquirencias
        adq_logs = [log for log in result.get('logs', []) if log.get('tipo') == 'adquirencia_cruzada']
        print(f"\n5. Cruces encontrados: {len(adq_logs)}")
        for i, log in enumerate(adq_logs[:10], 1):  # Mostrar primeros 10
            detalle = log.get('detalle', '')
            valor = log.get('valor', 0)
            fecha = log.get('fecha', '')
            print(f"   {i}. Valor: ${valor:,.2f} | Fecha: {fecha}")
            # Mostrar parte de autorización
            if 'aprobacion' in detalle:
                auth_part = detalle.split('aprobacion ')[-1].split(',')[0]
                print(f"      Auth: {auth_part}")
        
        if len(adq_logs) > 10:
            print(f"   ... y {len(adq_logs) - 10} más")
        
        # Obtener información de archivos generados
        print(f"\n6. Archivos generados:")
        for file_info in result.get('files', []):
            print(f"   - {file_info['name']}")
        
        # Descargar y verificar memorando procesado
        print(f"\n7. Verificando coloreo en Memorando procesado...")
        for file_info in result.get('files', []):
            if 'CRUCES' in file_info['name'].upper() or 'MEMORANDO' in file_info['name'].upper():
                try:
                    memorando_result_bytes = base64.b64decode(file_info['file'])
                    memorando_result_wb = load_workbook(BytesIO(memorando_result_bytes))
                    
                    # Buscar hoja con 690
                    ws_690 = None
                    for sheet_name in memorando_result_wb.sheetnames:
                        if '690' in sheet_name.lower():
                            ws_690 = memorando_result_wb[sheet_name]
                            break
                    
                    if ws_690:
                        # Contar celdas con color azul
                        blue_count = 0
                        blue_color = 'FFE8F4FF'
                        for row in ws_690.iter_rows(min_row=2, max_row=ws_690.max_row):
                            for cell in row:
                                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == blue_color:
                                    blue_count += 1
                        
                        print(f"   Hoja 690 encontrada")
                        print(f"   Celdas coloreadas AZUL: {blue_count}")
                        
                        # Buscar comentarios
                        comment_col = None
                        for col in range(1, ws_690.max_column + 1):
                            header = ws_690.cell(row=1, column=col).value
                            if header and 'coment' in str(header).lower():
                                comment_col = col
                                break
                        
                        if comment_col:
                            comments_with_adq = 0
                            for row in ws_690.iter_rows(min_row=2, max_row=ws_690.max_row):
                                cell = ws_690.cell(row=row[0].row, column=comment_col)
                                if cell.value and 'Adquirencia' in str(cell.value):
                                    comments_with_adq += 1
                            print(f"   Comentarios con 'Adquirencia': {comments_with_adq}")
                except Exception as e:
                    print(f"   Error verificando memorando: {e}")
        
        print(f"\n✓ PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        
    except Exception as e:
        print(f"✗ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
