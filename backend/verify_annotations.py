"""Verificar que los comentarios aparecen en AMBOS archivos."""

import base64
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from integrador import ProcesadorIntegrado

base = Path('..') / 'tmp_e2e'
adq = (base / 'ADQUIRENCIAS MAYO (1).xlsx').read_bytes()
memo = (base / 'CCS_Memorando Definitivo Ctas Bancarias Abril 2026.xlsx').read_bytes()

processor = ProcesadorIntegrado(
    contable_bytes=None,
    pse_bytes=None,
    cruces_bytes=memo,
    adquirencias_bytes=adq,
    date_tolerance_days=2,
    value_tolerance=0.01,
)

result = processor.procesar()

print("=== VERIFICANDO COMENTARIOS EN AMBOS ARCHIVOS ===\n")

# Verificar Memorando/Cruces
for file_info in result.get('files', []):
    if 'CRUCES' in file_info['name']:
        print(f"1. MEMORANDO ({file_info['name']}):")
        memorando_bytes = base64.b64decode(file_info['file'])
        memorando_wb = load_workbook(BytesIO(memorando_bytes))
        
        ws_690 = None
        for sheet_name in memorando_wb.sheetnames:
            if '690' in sheet_name.lower():
                ws_690 = memorando_wb[sheet_name]
                break
        
        if ws_690:
            comments_count = 0
            comment_col = None
            for col in range(1, ws_690.max_column + 1):
                header = ws_690.cell(1, col).value
                if header and isinstance(header, str) and 'coment' in str(header).lower():
                    comment_col = col
                    break
            
            if comment_col:
                for row in range(2, ws_690.max_row + 1):
                    cell = ws_690.cell(row, comment_col)
                    if cell.value and 'Adquirencia' in str(cell.value):
                        comments_count += 1
                print(f"   Hoja 690: {comments_count} comentarios con 'Adquirencia' (todas las filas)")
            else:
                print(f"   Hoja 690: No hay columna de comentarios")
    
    if 'ADQUIRENCIAS' in file_info['name']:
        print(f"\n2. ADQUIRENCIAS ({file_info['name']}):")
        adq_bytes = base64.b64decode(file_info['file'])
        adq_wb = load_workbook(BytesIO(adq_bytes))
        
        for sheet in adq_wb.worksheets:
            comments_count = 0
            comment_col = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(1, col).value
                if header and isinstance(header, str) and 'coment' in str(header).lower():
                    comment_col = col
                    break
            
            if comment_col:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row, comment_col)
                    if cell.value and 'Adquirencia' in str(cell.value):
                        comments_count += 1
                print(f"   Hoja {sheet.title}: {comments_count} comentarios con 'Adquirencia' (todas las filas)")

print("\n✓ VERIFICACIÓN COMPLETADA")
